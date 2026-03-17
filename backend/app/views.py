from __future__ import annotations

import ast
import json
import mimetypes
import os
import re
import uuid
import zipfile
from io import BytesIO
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote, urlparse
from xml.etree import ElementTree as ET

from django.conf import settings
from django.contrib.auth.hashers import check_password, make_password
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from sqlalchemy import func, select, text, update
from sqlalchemy.exc import IntegrityError

from .auth import login_with_username_password, require_auth
from .db import engine, get_session
from .models import (
    Customer,
    Item,
    ItemProductType,
    ItemTypeFormula,
    ItemTypeFormulaItem,
    ProductType,
    ProcessingPrice,
    Quotation,
    RawMaterialPrice,
    Material,
    MaterialCategory,
    MaterialGroup,
    FixedWeightTable,
    UnitWeightOption,
    Product,
    ProductPrintImage,
    ProductPrintVersion,
    ProductSpec,
    ProductionPlan,
    User,
    AuthToken,
)
from .utils import fmt_date, fmt_datetime, parse_date, to_num


def _body(request: HttpRequest):
    if not request.body:
        return {}
    return json.loads(request.body.decode("utf-8"))


def _str_or_none(v):
    if v is None:
        return None
    text = str(v).strip()
    return text if text else None


def _num_or_zero(v):
    if v is None:
        return 0
    if isinstance(v, str):
        raw = v.strip()
        if not raw:
            return 0
        raw = raw.replace(",", "")
        try:
            return float(raw)
        except ValueError:
            return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


def _norm_text(v: str | None) -> str:
    return (v or "").strip().lower()


def _norm_phone(v: str | None) -> str:
    return "".join((v or "").split()).lower()


def _upper_or_none(v):
    text = _str_or_none(v)
    return text.upper() if text else None


SPEC_ABC_PATTERN = re.compile(r"^\s*[^*]+\s*\*\s*\d+(?:\.\d+)?\s*\*\s*\d+(?:\.\d+)?\s*$")
FORMULA_ALLOWED_PATTERN = re.compile(r"^[A-Za-z0-9_+\-*/().\s]+$")
A_NUMBER_PATTERN = re.compile(r"[-+]?\d+(?:\.\d+)?")
PHI_TO_DIAMETER_PATTERN = re.compile(r"phi(?=\s*\d)|\bphi\b", re.IGNORECASE)
CUSTOMER_LEVEL_FACTORS = {
    "A": 0.97,
    "B": 0.98,
    "C": 1.00,
    "N": 1.02,
}
DEFAULT_PRODUCT_TYPES = [
    "BELT UPANEL",
    "ROPE UPANEL",
    "BELT TUBULAR",
    "ROPE TUBULAR",
    "4 PANEL",
    "BELT CIRCULAR",
    "ROPE CIRCULAR",
    "BAO CUỐN",
]


def _normalize_customer_level(v):
    level = _upper_or_none(v)
    if level is None:
        return None
    return level if level in CUSTOMER_LEVEL_FACTORS else None


def _infer_material_category_code_from_name(name: str | None) -> str:
    lowered = _norm_text(name)
    if any(x in lowered for x in {"dây", "day", "rope"}):
        return "ROPE"
    if any(x in lowered for x in {"vải", "vai", "fabric", "tarpaulin"}):
        return "FABRIC"
    return "OTHER"


def _normalize_material_category_code(code: str | None, name_for_fallback: str | None = None) -> str:
    raw = _str_or_none(code)
    if raw is None:
        return _infer_material_category_code_from_name(name_for_fallback)
    normalized = re.sub(r"[^A-Za-z0-9_]+", "_", raw).strip("_").upper()
    if not normalized:
        return _infer_material_category_code_from_name(name_for_fallback)
    return normalized


def _is_fabric_category(category_code: str | None, category_name: str | None) -> bool:
    code = _upper_or_none(category_code)
    if code in {"FABRIC", "VAI"}:
        return True
    lowered = _norm_text(category_name)
    return "vải" in lowered or "vai" in lowered


def _is_rope_category(category_code: str | None, category_name: str | None) -> bool:
    code = _upper_or_none(category_code)
    if code in {"ROPE", "DAY", "DAY_NHUA"}:
        return True
    lowered = _norm_text(category_name)
    return "dây" in lowered or "day" in lowered


def _normalize_lami_text(v):
    text = _str_or_none(v)
    if text is None:
        return None
    lowered = text.lower()
    if lowered in {"yes", "y", "true", "1"}:
        return "Yes"
    if lowered in {"no", "n", "false", "0"}:
        return "No"
    return None


def _normalize_phi_text(v):
    text = _str_or_none(v)
    if text is None:
        return None
    normalized = text.replace("Φ", "Ø").replace("ø", "Ø")
    normalized = PHI_TO_DIAMETER_PATTERN.sub("Ø", normalized)
    normalized = re.sub(r"Ø\s+(?=\d)", "Ø", normalized)
    return normalized


def _normalize_diameter_value(v):
    text = _str_or_none(v)
    if text is None:
        return None
    return re.sub(r"^\s*(?:phi|ø|Ø)\s*", "", text, flags=re.IGNORECASE)


def _normalize_spec_abc(v):
    text = _str_or_none(v)
    if text is None:
        return None
    if not SPEC_ABC_PATTERN.match(text):
        return None
    parts = [p.strip() for p in text.split("*")]
    parts[0] = _normalize_phi_text(parts[0]) or parts[0]
    return "*".join(parts)


def _normalize_number_text(v):
    text = _str_or_none(v)
    if text is None:
        return None
    normalized = text.replace(",", "")
    try:
        float(normalized)
        return normalized
    except (TypeError, ValueError):
        return None


def _parse_spec_abc(v: str | None):
    text = _str_or_none(v)
    if not text or not SPEC_ABC_PATTERN.match(text):
        return None
    parts = [p.strip() for p in text.split("*")]
    if len(parts) != 3:
        return None
    a_match = A_NUMBER_PATTERN.search(parts[0] or "")
    if a_match:
        try:
            a_num = float(a_match.group(0))
        except (TypeError, ValueError):
            a_num = None
    else:
        a_num = None
    try:
        b = float(parts[1])
        c = float(parts[2])
    except (TypeError, ValueError):
        return None
    return {"a_text": parts[0], "a_num": a_num, "b": b, "c": c}


def _parse_spec_parts(v: str | None, expected_parts: int):
    text = _str_or_none(v)
    if not text:
        return None
    parts = [p.strip() for p in text.split("*")]
    if len(parts) != expected_parts:
        return None
    values: list[float] = []
    for part in parts:
        m = A_NUMBER_PATTERN.search(part or "")
        if not m:
            return None
        try:
            values.append(float(m.group(0)))
        except (TypeError, ValueError):
            return None
    vars_map: dict[str, float] = {}
    if expected_parts >= 1:
        vars_map["A"] = values[0]
    if expected_parts >= 2:
        vars_map["B"] = values[1]
    if expected_parts >= 3:
        vars_map["C"] = values[2]
    return vars_map


def _split_pair_formula(expr: str | None):
    text = _str_or_none(expr)
    if text is None:
        return None
    # New format: left x right (x/X/× as separator at top level).
    # Keep backward compatibility for legacy formulas that used "*".
    depth = 0
    for i, ch in enumerate(text):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth < 0:
                return None
        elif ch in {"x", "X", "×"} and depth == 0:
            left = text[:i].strip()
            right = text[i + 1 :].strip()
            if not left or not right:
                return None
            return left, right
    depth = 0
    for i, ch in enumerate(text):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth < 0:
                return None
        elif ch == "*" and depth == 0:
            left = text[:i].strip()
            right = text[i + 1 :].strip()
            if not left or not right:
                return None
            return left, right
    return None


def _format_num_text(v: float):
    fixed = f"{v:.6f}"
    trimmed = fixed.rstrip("0").rstrip(".")
    return "0" if trimmed in {"", "-0"} else trimmed


def _normalize_ab_text(v: str | None):
    text = _str_or_none(v)
    if text is None:
        return None
    parts = [p.strip() for p in text.split("*")]
    if len(parts) != 2:
        return None
    a = A_NUMBER_PATTERN.search(parts[0] or "")
    b = A_NUMBER_PATTERN.search(parts[1] or "")
    if not a or not b:
        return None
    return f"{_format_num_text(float(a.group(0)))}*{_format_num_text(float(b.group(0)))}"


def _validate_formula_expr(expr: str | None) -> str | None:
    text = _str_or_none(expr)
    if text is None:
        return None
    text = re.sub(r"(?<=\d),(?=\d)", ".", text)
    if not FORMULA_ALLOWED_PATTERN.match(text):
        return None
    try:
        node = ast.parse(text, mode="eval")
    except Exception:
        return None

    allowed_nodes = (
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Name,
        ast.Load,
        ast.Constant,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.USub,
        ast.UAdd,
    )
    for child in ast.walk(node):
        if not isinstance(child, allowed_nodes):
            return None
        if isinstance(child, ast.Name) and child.id.upper() not in {"A", "B", "C"}:
            return None
        if isinstance(child, ast.Constant) and not isinstance(child.value, (int, float)):
            return None
    return text


def _evaluate_formula_expr(expr: str, vars_map: dict[str, float]):
    parsed = ast.parse(expr, mode="eval")

    def eval_node(node):
        if isinstance(node, ast.Expression):
            return eval_node(node.body)
        if isinstance(node, ast.Constant):
            return float(node.value)
        if isinstance(node, ast.Name):
            key = node.id.upper()
            if key not in vars_map:
                raise ValueError("Unknown variable")
            return float(vars_map[key])
        if isinstance(node, ast.UnaryOp):
            v = eval_node(node.operand)
            if isinstance(node.op, ast.USub):
                return -v
            if isinstance(node.op, ast.UAdd):
                return v
            raise ValueError("Unsupported unary op")
        if isinstance(node, ast.BinOp):
            left = eval_node(node.left)
            right = eval_node(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                if right == 0:
                    raise ZeroDivisionError("division by zero")
                return left / right
            raise ValueError("Unsupported bin op")
        raise ValueError("Unsupported expression")

    return eval_node(parsed)


def _compute_unit_weight(
    mode: str | None,
    fixed_value,
    formula_expr: str | None,
    spec_value: str | None,
    choice_value=None,
    use_lami_for_calc: bool = False,
    lami_calc_value=None,
):
    normalized_mode = (mode or "fixed").strip().lower()
    result = None
    if normalized_mode == "fixed":
        if fixed_value is None:
            return None
        try:
            result = float(fixed_value)
        except (TypeError, ValueError):
            return None

    elif normalized_mode == "choice":
        if choice_value is None:
            return None
        try:
            result = float(choice_value)
        except (TypeError, ValueError):
            return None

    elif normalized_mode != "formula":
        return None
    else:
        expr = _validate_formula_expr(formula_expr)
        if not expr:
            return None
        parsed = _parse_spec_abc(spec_value)
        if not parsed:
            return None
        vars_map = {
            "B": parsed["b"],
            "C": parsed["c"],
        }
        if parsed["a_num"] is not None:
            vars_map["A"] = parsed["a_num"]
        try:
            result = float(_evaluate_formula_expr(expr, vars_map))
        except Exception:
            return None

    if use_lami_for_calc:
        try:
            lami_value = float(lami_calc_value)
        except (TypeError, ValueError):
            return None
        result = result + lami_value
    return result


def _compute_item_size(
    mode: str | None,
    fixed_type: str | None,
    fixed_value,
    fixed_text: str | None,
    formula_expr: str | None,
    source_value: str | None,
    source_field: str | None = None,
):
    normalized_mode = (mode or "fixed").strip().lower()
    if normalized_mode == "fixed":
        normalized_fixed_type = (fixed_type or "number").strip().lower()
        if normalized_fixed_type == "ab":
            return _normalize_ab_text(fixed_text)
        if fixed_value is None:
            return None
        try:
            return _format_num_text(float(fixed_value))
        except (TypeError, ValueError):
            return None
    if normalized_mode != "formula":
        return None
    source = (source_field or "spec_inner").strip().lower()
    if source == "liner" and not _str_or_none(formula_expr):
        return _str_or_none(source_value)
    pair = _split_pair_formula(formula_expr)
    if not pair:
        return None
    left_expr, right_expr = pair
    left_expr = _validate_formula_expr(left_expr)
    right_expr = _validate_formula_expr(right_expr)
    if not left_expr or not right_expr:
        return None
    vars_map: dict[str, float] | None = None
    if source == "spec_inner":
        vars_map = _parse_spec_parts(source_value, 3) or _parse_spec_parts(source_value, 2)
    elif source == "liner":
        vars_map = _parse_spec_parts(source_value, 3)
    elif source in {"top", "bottom"}:
        vars_map = _parse_spec_parts(source_value, 2)
    if vars_map is None:
        text = _str_or_none(source_value)
        if not text:
            return None
        first_num = A_NUMBER_PATTERN.search(text)
        if not first_num:
            return None
        try:
            vars_map = {"A": float(first_num.group(0))}
        except (TypeError, ValueError):
            return None
    try:
        left = float(_evaluate_formula_expr(left_expr, vars_map))
        right = float(_evaluate_formula_expr(right_expr, vars_map))
        return f"{_format_num_text(left)}*{_format_num_text(right)}"
    except Exception:
        return None


def _resolve_item_size_source_value(source_field: str | None, product: Product, fallback_spec: str | None):
    source = (source_field or "spec_inner").strip().lower()
    if source == "spec_inner":
        return _str_or_none(product.spec_inner) or _str_or_none(fallback_spec)
    if source == "top":
        return _str_or_none(product.top)
    if source == "bottom":
        return _str_or_none(product.bottom)
    if source == "liner":
        return _str_or_none(product.liner)
    return _str_or_none(product.spec_inner) or _str_or_none(fallback_spec)


def _compute_item_size_by_product_type_formula(
    session,
    product: Product | None,
    item: Item | None,
    fallback_spec: str | None = None,
):
    if item is None or product is None:
        return None
    source_field = _str_or_none(item.item_size_source_field) or "spec_inner"
    source_value = _resolve_item_size_source_value(source_field, product, fallback_spec)
    product_type_name = _upper_or_none(product.type)
    product_type_formula = None
    if product_type_name:
        pt = session.scalar(
            select(ProductType).where(
                ProductType.product_type_name == product_type_name,
                _active(ProductType),
            )
        )
        if pt is not None:
            row = session.scalar(
                select(ItemTypeFormula).where(
                    ItemTypeFormula.item_id == item.id,
                    ItemTypeFormula.product_type_id == pt.id,
                )
            )
            if row is not None:
                product_type_formula = _str_or_none(row.formula)
    if not product_type_formula:
        if (source_field or "").strip().lower() == "liner":
            return _compute_item_size(
                "formula",
                None,
                None,
                None,
                None,
                source_value,
                source_field,
            )
        # Strict rule: without Product Type formula, item size is not computed.
        return None
    return _compute_item_size(
        "formula",
        None,
        None,
        None,
        product_type_formula,
        source_value,
        source_field,
    )


def _compute_qty_from_item_size(item_size_value):
    if item_size_value is None:
        return None
    if isinstance(item_size_value, (int, float)):
        return float(item_size_value)
    text = _str_or_none(item_size_value)
    if not text:
        return None
    parts = [p.strip() for p in text.split("*")]
    nums: list[float] = []
    for p in parts:
        m = A_NUMBER_PATTERN.search(p or "")
        if not m:
            return None
        try:
            nums.append(float(m.group(0)))
        except (TypeError, ValueError):
            return None
    if not nums:
        return None
    out = 1.0
    for n in nums:
        out *= n
    # Item Size in form A*B is in cm, convert to m2.
    if len(nums) == 2:
        out = out / 10000.0
    return out


def _to_float_or_none(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    raw = _str_or_none(v)
    if raw is None:
        return None
    try:
        return float(raw.replace(",", ""))
    except (TypeError, ValueError):
        return None


def _compute_wt_kg(unit_weight_kg, qty_m_or_m2, pcs_ea):
    unit_weight = _to_float_or_none(unit_weight_kg)
    qty = _to_float_or_none(qty_m_or_m2)
    pcs = _to_float_or_none(pcs_ea)
    if unit_weight is None or qty is None or pcs is None:
        return None
    return unit_weight * qty * pcs


def _compute_unit_weight_from_item_material(session, item: Item | None, spec_value: str | None, lami_value: str | None = None):
    if item is None or not item.material_id:
        return None, None, None
    material = session.scalar(select(Material).where(Material.id == item.material_id, _active(Material)))
    if material is None:
        return None, None, None
    category_name = material.material_category.material_category_name if material.material_category else None
    category_code = material.material_category.material_category_code if material.material_category else None

    if _is_fabric_category(category_code, category_name):
        resolved_spec = _str_or_none(spec_value)
        if not resolved_spec:
            return None, material, None
        normalized_lami = _normalize_lami_text(lami_value)
        if normalized_lami is None:
            normalized_lami = "Yes" if bool(material.lami) else "No"
        computed = _compute_unit_weight(
            "formula",
            None,
            material.formula,
            resolved_spec,
            None,
            normalized_lami == "Yes",
            0.025 if normalized_lami == "Yes" else None,
        )
        return computed, material, resolved_spec

    if _is_rope_category(category_code, category_name):
        chosen_spec = _str_or_none(spec_value)
        fixed_row = None
        if chosen_spec:
            fixed_row = session.scalar(
                select(FixedWeightTable).where(
                    FixedWeightTable.material_id == material.id,
                    FixedWeightTable.size_label == chosen_spec,
                    _active(FixedWeightTable),
                )
            )
        if not fixed_row:
            fixed_row = session.scalar(
                select(FixedWeightTable)
                .where(FixedWeightTable.material_id == material.id, _active(FixedWeightTable))
                .order_by(FixedWeightTable.id.asc())
            )
        if not fixed_row:
            return None, material, chosen_spec
        return _to_float_or_none(fixed_row.unit_weight_value), material, fixed_row.size_label

    resolved_spec = _str_or_none(spec_value)
    return None, material, resolved_spec


def _sync_product_spec_from_item_material(session, spec_row: ProductSpec):
    item = session.scalar(select(Item).where(Item.id == spec_row.item_id, _active(Item)))
    if not item or not item.material_id:
        return
    product = session.scalar(select(Product).where(Product.id == spec_row.product_id, _active(Product)))
    if product:
        auto_item_size = _compute_item_size_by_product_type_formula(session, product, item, spec_row.spec)
        if auto_item_size:
            spec_row.item_size = auto_item_size
    computed_unit_weight, material, resolved_spec = _compute_unit_weight_from_item_material(
        session,
        item,
        spec_row.spec,
        spec_row.lami,
    )
    if material is None:
        return
    category_name = material.material_category.material_category_name if material.material_category else None
    category_code = material.material_category.material_category_code if material.material_category else None
    if _is_fabric_category(category_code, category_name):
        if not _str_or_none(spec_row.spec) and resolved_spec:
            spec_row.spec = resolved_spec
        if _normalize_lami_text(spec_row.lami) is None:
            spec_row.lami = "Yes" if bool(material.lami) else "No"
    elif _is_rope_category(category_code, category_name):
        if resolved_spec:
            spec_row.spec = resolved_spec
        spec_row.lami = "No"
    else:
        if not _str_or_none(spec_row.spec) and resolved_spec:
            spec_row.spec = resolved_spec
        if _normalize_lami_text(spec_row.lami) is None:
            spec_row.lami = "No"
    if computed_unit_weight is not None:
        spec_row.unit_weight_kg = computed_unit_weight
    spec_row.wt_kg = _compute_wt_kg(spec_row.unit_weight_kg, spec_row.qty_m_or_m2, spec_row.pcs_ea)
    session.add(spec_row)


def _next_product_spec_line_no(session, product_id: int) -> int:
    # Keep line_no unique within a product across all rows, including soft-deleted rows.
    max_line_no = session.scalar(select(func.max(ProductSpec.line_no)).where(ProductSpec.product_id == product_id)) or 0
    return int(max_line_no) + 1


def _recompute_product_specs_item_size_qty(session, product: Product):
    specs = session.scalars(
        select(ProductSpec).where(ProductSpec.product_id == product.id, _active(ProductSpec)).order_by(ProductSpec.id.asc())
    ).all()
    if not specs:
        return
    for spec in specs:
        item = session.scalar(select(Item).where(Item.id == spec.item_id, _active(Item)))
        auto_item_size = _compute_item_size_by_product_type_formula(session, product, item, spec.spec)
        if auto_item_size:
            spec.item_size = auto_item_size
        spec.qty_m_or_m2 = _compute_qty_from_item_size(spec.item_size)
        spec.wt_kg = _compute_wt_kg(spec.unit_weight_kg, spec.qty_m_or_m2, spec.pcs_ea)
        session.add(spec)


def _xlsx_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    path = "xl/sharedStrings.xml"
    if path not in zf.namelist():
        return []
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(zf.read(path))
    out: list[str] = []
    for si in root.findall("a:si", ns):
        out.append("".join(t.text or "" for t in si.findall(".//a:t", ns)))
    return out


def _xlsx_sheet_rows(file_path: str, sheet_name: str) -> list[list[str | None]]:
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(file_path) as zf:
        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship")
        }
        shared_strings = _xlsx_shared_strings(zf)

        target = None
        for sh in wb.findall("a:sheets/a:sheet", ns):
            if sh.attrib.get("name") == sheet_name:
                rid = sh.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                rel_target = rel_map[rid]
                target = rel_target if rel_target.startswith("xl/") else f"xl/{rel_target.lstrip('/')}"
                break
        if not target:
            raise ValueError(f"Không tìm thấy sheet '{sheet_name}'")

        ws = ET.fromstring(zf.read(target))
        rows: list[list[str | None]] = []
        for row in ws.findall("a:sheetData/a:row", ns):
            cells: dict[int, str | None] = {}
            max_col = 0
            for cell in row.findall("a:c", ns):
                ref = cell.attrib.get("r", "")
                col_letters = "".join(ch for ch in ref if ch.isalpha())
                col_index = 0
                for ch in col_letters:
                    col_index = col_index * 26 + (ord(ch.upper()) - ord("A") + 1)
                if col_index <= 0:
                    continue
                max_col = max(max_col, col_index)

                cell_type = cell.attrib.get("t")
                if cell_type == "inlineStr":
                    is_node = cell.find("a:is", ns)
                    value = "".join(t.text or "" for t in is_node.findall(".//a:t", ns)) if is_node is not None else None
                else:
                    v = cell.find("a:v", ns)
                    value = v.text if v is not None else None
                    if cell_type == "s" and value is not None and value.isdigit():
                        idx = int(value)
                        value = shared_strings[idx] if 0 <= idx < len(shared_strings) else value
                cells[col_index] = value

            if max_col == 0:
                rows.append([])
                continue
            rows.append([cells.get(i) for i in range(1, max_col + 1)])
        return rows


def _ok(data, status=200):
    return JsonResponse(data, status=status, safe=False)


def frontend_app(request: HttpRequest, path: str = ""):
    dist_dir = Path(settings.FRONTEND_DIST_DIR)
    if not dist_dir.exists():
        return HttpResponse("Frontend dist not found", status=404)

    safe_path = (path or "").lstrip("/")
    requested = (dist_dir / safe_path).resolve()
    try:
        requested.relative_to(dist_dir.resolve())
    except ValueError:
        return HttpResponse("Invalid path", status=400)

    target = requested if safe_path and requested.exists() and requested.is_file() else (dist_dir / "index.html")
    if not target.exists():
        return HttpResponse("index.html not found", status=404)

    ctype, _ = mimetypes.guess_type(str(target))
    with target.open("rb") as fh:
        return HttpResponse(fh.read(), content_type=ctype or "text/html")


def _now() -> datetime:
    return datetime.now()


def _active(model):
    return model.deleted_at.is_(None)


def _is_admin(user: User) -> bool:
    return user.role == "admin"


def _forbidden():
    return _ok({"detail": "Bạn không có quyền thực hiện thao tác này"}, 403)


def _ensure_production_plan_note_column() -> None:
    with engine.begin() as conn:
        cols = {row[1] for row in conn.execute(text("PRAGMA table_info(production_plans)")).fetchall()}
        if "note" not in cols:
            conn.execute(text("ALTER TABLE production_plans ADD COLUMN note TEXT"))


def _ensure_raw_material_prices_table() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS raw_material_prices (
                    id INTEGER NOT NULL PRIMARY KEY,
                    material_name VARCHAR(100) NOT NULL UNIQUE,
                    unit VARCHAR(30) NOT NULL DEFAULT 'kg',
                    unit_price NUMERIC(12,4) NOT NULL DEFAULT 0,
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_raw_material_prices_material_name ON raw_material_prices (material_name)"))
        defaults = [
            ("TB", "kg", 0.85),
            ("PP", "kg", 1.00),
            ("HDPE", "kg", 1.10),
            ("LDPE", "kg", 1.40),
            ("LLDPE", "kg", 1.50),
        ]
        for name, unit, price in defaults:
            conn.execute(
                text(
                    """
                    INSERT INTO raw_material_prices (material_name, unit, unit_price, created_at, updated_at)
                    SELECT :name, :unit, :price, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                    WHERE NOT EXISTS (
                        SELECT 1 FROM raw_material_prices WHERE material_name = :name
                    )
                    """
                ),
                {"name": name, "unit": unit, "price": price},
            )


def _ensure_processing_prices_table() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS processing_prices (
                    id INTEGER NOT NULL PRIMARY KEY,
                    process_name VARCHAR(100) NOT NULL UNIQUE,
                    unit_price NUMERIC(12,4) NOT NULL DEFAULT 0,
                    note TEXT,
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )


def _ensure_product_types_table() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS product_types (
                    id INTEGER NOT NULL PRIMARY KEY,
                    product_type_name VARCHAR(100) NOT NULL UNIQUE,
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        cols = [row[1] for row in conn.execute(text("PRAGMA table_info(product_types)")).fetchall()]
        if "formula" in cols:
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS product_types__new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        product_type_name VARCHAR(100) NOT NULL UNIQUE,
                        deleted_at DATETIME,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                    )
                    """
                )
            )
            conn.execute(text("DELETE FROM product_types__new"))
            conn.execute(
                text(
                    """
                    INSERT INTO product_types__new (id, product_type_name, deleted_at, created_at, updated_at)
                    SELECT id, product_type_name, deleted_at, created_at, updated_at
                    FROM product_types
                    """
                )
            )
            conn.execute(text("DROP TABLE product_types"))
            conn.execute(text("ALTER TABLE product_types__new RENAME TO product_types"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_product_types_product_type_name ON product_types (product_type_name)"))
        for name in DEFAULT_PRODUCT_TYPES:
            existing = conn.execute(
                text("SELECT id, deleted_at FROM product_types WHERE product_type_name = :name"),
                {"name": name},
            ).fetchone()
            if existing and existing[1] is not None:
                conn.execute(
                    text(
                        """
                        UPDATE product_types
                        SET deleted_at = NULL, updated_at = CURRENT_TIMESTAMP
                        WHERE id = :id
                        """
                    ),
                    {"id": existing[0]},
                )
            if not existing:
                conn.execute(
                    text(
                        """
                        INSERT INTO product_types (product_type_name, created_at, updated_at)
                        VALUES (:name, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                        """
                    ),
                    {"name": name},
                )
        conn.execute(
            text(
                """
                UPDATE product_types
                SET deleted_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
                WHERE product_type_name = 'OTHER' AND deleted_at IS NULL
                """
            )
        )


def _ensure_quotations_table() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS quotations (
                    id INTEGER NOT NULL PRIMARY KEY,
                    customer_id INTEGER NOT NULL,
                    product_id INTEGER NOT NULL,
                    has_lami BOOLEAN NOT NULL DEFAULT 0,
                    lami_unit_price NUMERIC(12,4) NOT NULL DEFAULT 0.429,
                    level_code VARCHAR(10),
                    level_factor NUMERIC(12,4) NOT NULL DEFAULT 1,
                    size_value VARCHAR(255),
                    total_weight_kg NUMERIC(14,4) NOT NULL DEFAULT 0,
                    pe_weight_kg NUMERIC(14,4) NOT NULL DEFAULT 0,
                    pp_weight_kg NUMERIC(14,4) NOT NULL DEFAULT 0,
                    amount_weight NUMERIC(14,4) NOT NULL DEFAULT 0,
                    amount_lami NUMERIC(14,4) NOT NULL DEFAULT 0,
                    amount_color NUMERIC(14,4) NOT NULL DEFAULT 0,
                    amount_extra NUMERIC(14,4) NOT NULL DEFAULT 0,
                    subtotal NUMERIC(14,4) NOT NULL DEFAULT 0,
                    total NUMERIC(14,4) NOT NULL DEFAULT 0,
                    row_payload TEXT,
                    note TEXT,
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_quotations_customer_id ON quotations (customer_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_quotations_product_id ON quotations (product_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_quotations_level_code ON quotations (level_code)"))


def _ensure_fixed_weight_tables_table() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS fixed_weight_tables (
                    id INTEGER NOT NULL PRIMARY KEY,
                    material_id INTEGER,
                    size_label VARCHAR(100) NOT NULL,
                    unit_weight_value NUMERIC(12,5) NOT NULL DEFAULT 0,
                    unit_price NUMERIC(12,4) NOT NULL DEFAULT 0,
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        cols = {row[1] for row in conn.execute(text("PRAGMA table_info(fixed_weight_tables)")).fetchall()}
        if "material_group_id" in cols:
            conn.execute(text("PRAGMA foreign_keys=OFF"))
            conn.execute(text("DROP TABLE IF EXISTS fixed_weight_tables__new"))
            conn.execute(
                text(
                    """
                    CREATE TABLE fixed_weight_tables__new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        material_id INTEGER,
                        size_label VARCHAR(100) NOT NULL,
                        unit_weight_value NUMERIC(12,5) NOT NULL DEFAULT 0,
                        unit_price NUMERIC(12,4) NOT NULL DEFAULT 0,
                        deleted_at DATETIME,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    INSERT INTO fixed_weight_tables__new (id, material_id, size_label, unit_weight_value, unit_price, deleted_at, created_at, updated_at)
                    SELECT
                        f.id,
                        (
                            SELECT m.id
                            FROM material_groups mg
                            JOIN materials m ON TRIM(LOWER(m.material_name)) = TRIM(LOWER(mg.material_group_name))
                            WHERE mg.id = f.material_group_id
                            LIMIT 1
                        ) AS material_id,
                        f.size_label,
                        f.unit_weight_value,
                        f.unit_price,
                        f.deleted_at,
                        COALESCE(f.created_at, CURRENT_TIMESTAMP),
                        COALESCE(f.updated_at, CURRENT_TIMESTAMP)
                    FROM fixed_weight_tables f
                    """
                )
            )
            conn.execute(text("DROP TABLE fixed_weight_tables"))
            conn.execute(text("ALTER TABLE fixed_weight_tables__new RENAME TO fixed_weight_tables"))
            conn.execute(text("PRAGMA foreign_keys=ON"))
        cols_after = {row[1] for row in conn.execute(text("PRAGMA table_info(fixed_weight_tables)")).fetchall()}
        if "material_id" not in cols_after:
            conn.execute(text("ALTER TABLE fixed_weight_tables ADD COLUMN material_id INTEGER"))
        conn.execute(text("DROP INDEX IF EXISTS ix_fixed_weight_tables_material_group_id"))
        conn.execute(text("DROP INDEX IF EXISTS uq_fwt_mg_size_active"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_fixed_weight_tables_material_id ON fixed_weight_tables (material_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_fixed_weight_tables_size_label ON fixed_weight_tables (size_label)"))
        conn.execute(
            text(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS uq_fwt_material_size_active
                ON fixed_weight_tables(material_id, size_label)
                WHERE deleted_at IS NULL AND material_id IS NOT NULL
                """
            )
        )


def _ensure_material_master_tables() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS material_categories (
                    id INTEGER NOT NULL PRIMARY KEY,
                    material_category_name VARCHAR(100) NOT NULL UNIQUE,
                    material_category_code VARCHAR(50),
                    spec_format VARCHAR(20) NOT NULL DEFAULT 'text',
                    format VARCHAR(255),
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS materials (
                    id INTEGER NOT NULL PRIMARY KEY,
                    material_name VARCHAR(100) NOT NULL UNIQUE,
                    material_category_id INTEGER,
                    formula VARCHAR(255),
                    spec VARCHAR(255),
                    lami BOOLEAN NOT NULL DEFAULT 0,
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_material_categories_name ON material_categories (material_category_name)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_materials_name ON materials (material_name)"))
        mat_cols = {row[1] for row in conn.execute(text("PRAGMA table_info(materials)")).fetchall()}
        if "material_category_id" not in mat_cols:
            conn.execute(text("ALTER TABLE materials ADD COLUMN material_category_id INTEGER"))
        if "formula" not in mat_cols:
            conn.execute(text("ALTER TABLE materials ADD COLUMN formula VARCHAR(255)"))
        if "spec" not in mat_cols:
            conn.execute(text("ALTER TABLE materials ADD COLUMN spec VARCHAR(255)"))
        if "lami" not in mat_cols:
            conn.execute(text("ALTER TABLE materials ADD COLUMN lami BOOLEAN NOT NULL DEFAULT 0"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_materials_material_category_id ON materials (material_category_id)"))
        mc_cols = {row[1] for row in conn.execute(text("PRAGMA table_info(material_categories)")).fetchall()}
        if "material_category_code" not in mc_cols:
            conn.execute(text("ALTER TABLE material_categories ADD COLUMN material_category_code VARCHAR(50)"))
        if "spec_format" not in mc_cols:
            conn.execute(text("ALTER TABLE material_categories ADD COLUMN spec_format VARCHAR(20) NOT NULL DEFAULT 'text'"))
        if "format" not in mc_cols:
            conn.execute(text('ALTER TABLE material_categories ADD COLUMN "format" VARCHAR(255)'))
        if "formula" in mc_cols:
            conn.execute(text('UPDATE material_categories SET "format" = formula WHERE ("format" IS NULL OR TRIM("format") = "") AND formula IS NOT NULL'))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_material_categories_code ON material_categories (material_category_code)"))
        rows = conn.execute(text("SELECT id, material_category_name, material_category_code FROM material_categories")).fetchall()
        for row_id, row_name, row_code in rows:
            if _str_or_none(row_code):
                continue
            inferred = _normalize_material_category_code(None, row_name)
            conn.execute(
                text(
                    """
                    UPDATE material_categories
                    SET material_category_code = :code,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = :id
                    """
                ),
                {"id": row_id, "code": inferred},
            )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_processing_prices_process_name ON processing_prices (process_name)"))
        conn.execute(
            text(
                """
                INSERT INTO processing_prices (process_name, unit_price, note, created_at, updated_at)
                SELECT 'Gia công', 1.0, NULL, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                WHERE NOT EXISTS (
                    SELECT 1 FROM processing_prices WHERE process_name = 'Gia công'
                )
                """
            )
        )


def _ensure_items_table_schema() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS items (
                    id INTEGER NOT NULL PRIMARY KEY,
                    item_name VARCHAR(255) NOT NULL UNIQUE,
                    material_id INTEGER,
                    item_size_source_field VARCHAR(20),
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        cols = {row[1] for row in conn.execute(text("PRAGMA table_info(items)")).fetchall()}
        legacy_cols = {
            "item_color",
            "item_size_mode",
            "item_size_fixed_type",
            "item_size_value",
            "item_size_value_text",
            "item_size_formula_code",
        }
        if cols and legacy_cols.intersection(cols):
            conn.execute(text("PRAGMA foreign_keys=OFF"))
            conn.execute(text("DROP TABLE IF EXISTS items__new"))
            conn.execute(
                text(
                    """
                    CREATE TABLE items__new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        item_name VARCHAR(255) NOT NULL UNIQUE,
                        material_id INTEGER,
                        item_size_source_field VARCHAR(20),
                        deleted_at DATETIME,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    INSERT INTO items__new (id, item_name, material_id, item_size_source_field, deleted_at, created_at, updated_at)
                    SELECT id, item_name, NULL, 'spec_inner', deleted_at, COALESCE(created_at, CURRENT_TIMESTAMP), COALESCE(updated_at, CURRENT_TIMESTAMP)
                    FROM items
                    """
                )
            )
            conn.execute(text("DROP TABLE items"))
            conn.execute(text("ALTER TABLE items__new RENAME TO items"))
            conn.execute(text("PRAGMA foreign_keys=ON"))
        cols_after = {row[1] for row in conn.execute(text("PRAGMA table_info(items)")).fetchall()}
        if "material_id" not in cols_after:
            conn.execute(text("ALTER TABLE items ADD COLUMN material_id INTEGER"))
        if "item_size_source_field" not in cols_after:
            conn.execute(text("ALTER TABLE items ADD COLUMN item_size_source_field VARCHAR(20)"))
        conn.execute(
            text(
                """
                UPDATE items
                SET item_size_source_field = 'spec_inner'
                WHERE item_size_source_field IS NULL OR TRIM(item_size_source_field) = ''
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_items_item_name ON items (item_name)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_items_material_id ON items (material_id)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS item_product_types (
                    item_id INTEGER NOT NULL,
                    product_type_id INTEGER NOT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY (item_id, product_type_id)
                )
                """
            )
        )
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_item_product_types_item_id ON item_product_types (item_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_item_product_types_product_type_id ON item_product_types (product_type_id)"))
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS item_type_formulas (
                    item_id INTEGER NOT NULL,
                    product_type_id INTEGER NOT NULL,
                    formula TEXT NOT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    PRIMARY KEY (item_id, product_type_id)
                )
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS item_type_formula_items (
                    item_id INTEGER NOT NULL PRIMARY KEY,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        # Keep formula matrix empty by default.
        # Backward compatibility cleanup: only reset when table looks exactly like
        # legacy auto-seed (all active items copied and no formulas yet).
        seeded = int(conn.execute(text("SELECT COUNT(1) FROM item_type_formula_items")).scalar() or 0)
        formula_count = int(conn.execute(text("SELECT COUNT(1) FROM item_type_formulas")).scalar() or 0)
        active_items = int(conn.execute(text("SELECT COUNT(1) FROM items WHERE deleted_at IS NULL")).scalar() or 0)
        if seeded > 0 and formula_count == 0 and seeded == active_items:
            conn.execute(text("DELETE FROM item_type_formula_items"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_item_type_formulas_item_id ON item_type_formulas (item_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_item_type_formulas_product_type_id ON item_type_formulas (product_type_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_item_type_formula_items_item_id ON item_type_formula_items (item_id)"))


def _ensure_product_specs_schema() -> None:
    with engine.begin() as conn:
        cols = conn.execute(text("PRAGMA table_info(product_specs)")).fetchall()
        if not cols:
            return
        col_map = {row[1]: row for row in cols}
        mg_col = col_map.get("material_group_id")
        # Rebuild table only when material_group_id is NOT NULL (legacy schema).
        if mg_col is None or int(mg_col[3] or 0) == 0:
            return
        conn.execute(text("PRAGMA foreign_keys=OFF"))
        conn.execute(text("DROP TABLE IF EXISTS product_specs__new"))
        conn.execute(
            text(
                """
                CREATE TABLE product_specs__new (
                    id INTEGER NOT NULL PRIMARY KEY,
                    product_id INTEGER NOT NULL,
                    item_id INTEGER NOT NULL,
                    material_group_id INTEGER,
                    line_no INTEGER NOT NULL DEFAULT 1,
                    spec VARCHAR(255),
                    item_size VARCHAR(100),
                    lami VARCHAR(100),
                    item_color VARCHAR(100),
                    unit_weight_kg NUMERIC(12,4),
                    qty_m_or_m2 NUMERIC(14,4),
                    pcs_ea NUMERIC(12,4),
                    wt_kg NUMERIC(14,4),
                    other_note TEXT,
                    is_manual_weight BOOLEAN NOT NULL DEFAULT 1,
                    deleted_at DATETIME,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        conn.execute(
            text(
                """
                INSERT INTO product_specs__new (
                    id, product_id, item_id, material_group_id, line_no, spec, item_size, lami, item_color,
                    unit_weight_kg, qty_m_or_m2, pcs_ea, wt_kg, other_note, is_manual_weight, deleted_at, created_at, updated_at
                )
                SELECT
                    id, product_id, item_id, material_group_id, line_no, spec, item_size, lami, item_color,
                    unit_weight_kg, qty_m_or_m2, pcs_ea, wt_kg, other_note, COALESCE(is_manual_weight, 1), deleted_at,
                    COALESCE(created_at, CURRENT_TIMESTAMP), COALESCE(updated_at, CURRENT_TIMESTAMP)
                FROM product_specs
                """
            )
        )
        conn.execute(text("DROP TABLE product_specs"))
        conn.execute(text("ALTER TABLE product_specs__new RENAME TO product_specs"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_product_specs_product_id ON product_specs (product_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_product_specs_item_id ON product_specs (item_id)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_product_specs_material_group_id ON product_specs (material_group_id)"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_productspec_product_line ON product_specs (product_id, line_no)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_ps_product_line ON product_specs (product_id, line_no)"))
        conn.execute(text("PRAGMA foreign_keys=ON"))


def serialize_user(u: User):
    return {
        "id": u.id,
        "username": u.username,
        "full_name": u.full_name,
        "avatar_url": u.avatar_url,
        "role": u.role,
        "is_active": u.is_active,
        "created_at": fmt_datetime(u.created_at),
        "updated_at": fmt_datetime(u.updated_at),
    }


def serialize_customer(c: Customer):
    level = _normalize_customer_level(c.level)
    level_factor = CUSTOMER_LEVEL_FACTORS.get(level) if level else None
    return {
        "id": c.id,
        "customer_code": c.customer_code,
        "customer_name": c.customer_name,
        "address": c.address,
        "contact_person": c.contact_person,
        "phone": c.phone,
        "email": c.email,
        "production_2025": to_num(c.production_2025),
        "production_2026": to_num(c.production_2026),
        "in_production": to_num(c.in_production),
        "level": level,
        "level_factor": level_factor,
        "level_percent": (level_factor * 100.0) if level_factor is not None else None,
        "created_at": fmt_datetime(c.created_at),
        "updated_at": fmt_datetime(c.updated_at),
    }


def serialize_product(p: Product):
    return {
        "id": p.id,
        "customer_id": p.customer_id,
        "product_code": p.product_code,
        "product_name": p.product_name,
        "swl": p.swl,
        "type": p.type,
        "sewing_type": p.sewing_type,
        "print": p.print,
        "spec_other": p.spec_other,
        "spec_inner": p.spec_inner,
        "color": p.color,
        "liner": p.liner,
        "has_print_assets": p.has_print_assets,
        "top": p.top,
        "bottom": p.bottom,
        "packing": p.packing,
        "other_note": p.other_note,
        "created_at": fmt_datetime(p.created_at),
        "updated_at": fmt_datetime(p.updated_at),
    }


def serialize_product_type(p: ProductType):
    return {
        "id": p.id,
        "product_type_name": p.product_type_name,
        "created_at": fmt_datetime(p.created_at),
        "updated_at": fmt_datetime(p.updated_at),
    }


def _get_item_product_type_map(session, item_ids: list[int]):
    mapping: dict[int, dict[str, list]] = {item_id: {"ids": [], "names": []} for item_id in item_ids}
    if not item_ids:
        return mapping
    rows = session.execute(
        select(ItemProductType.item_id, ProductType.id, ProductType.product_type_name)
        .join(ProductType, ProductType.id == ItemProductType.product_type_id)
        .where(
            ItemProductType.item_id.in_(item_ids),
            _active(ProductType),
            ProductType.product_type_name != "OTHER",
        )
    ).all()
    for item_id, product_type_id, product_type_name in rows:
        entry = mapping.setdefault(item_id, {"ids": [], "names": []})
        entry["ids"].append(int(product_type_id))
        entry["names"].append(product_type_name)
    return mapping


def serialize_item(i: Item, product_type_map: dict[int, dict[str, list]] | None = None):
    mapped = (product_type_map or {}).get(i.id, {"ids": [], "names": []})
    return {
        "id": i.id,
        "item_name": i.item_name,
        "material_id": i.material_id,
        "material_name": i.material.material_name if i.material else None,
        "product_type_ids": mapped.get("ids", []),
        "product_type_names": mapped.get("names", []),
        "item_size_mode": None,
        "item_size_fixed_type": None,
        "item_size_value": None,
        "item_size_value_text": None,
        "item_size_formula_code": None,
        "item_size_formula": None,
        "item_size_source_field": i.item_size_source_field or "spec_inner",
        "created_at": fmt_datetime(i.created_at),
        "updated_at": fmt_datetime(i.updated_at),
    }


def serialize_raw_material_price(r: RawMaterialPrice):
    return {
        "id": r.id,
        "material_name": r.material_name,
        "unit": r.unit,
        "unit_price": to_num(r.unit_price),
        "created_at": fmt_datetime(r.created_at),
        "updated_at": fmt_datetime(r.updated_at),
    }


def serialize_processing_price(r: ProcessingPrice):
    return {
        "id": r.id,
        "process_name": r.process_name,
        "unit_price": to_num(r.unit_price),
        "note": r.note,
        "created_at": fmt_datetime(r.created_at),
        "updated_at": fmt_datetime(r.updated_at),
    }


def _sanitize_extra_rows(rows) -> list[dict]:
    out: list[dict] = []
    if not isinstance(rows, list):
        return out
    for row in rows:
        if not isinstance(row, dict):
            continue
        name = _str_or_none(row.get("name")) or ""
        value = _str_or_none(row.get("value"))
        amount = _to_float_or_none(row.get("amount"))
        out.append(
            {
                "name": name,
                "value": value,
                "amount": 0.0 if amount is None else amount,
            }
        )
    return out


def _sum_product_spec_weight(session, product_id: int):
    specs = session.scalars(
        select(ProductSpec)
        .where(ProductSpec.product_id == product_id, _active(ProductSpec))
        .order_by(ProductSpec.id.asc())
    ).all()
    total_weight = 0.0
    pe_weight = 0.0
    pe_names = {"tarpaulin", "pe liner"}
    for spec in specs:
        wt = _compute_wt_kg(spec.unit_weight_kg, spec.qty_m_or_m2, spec.pcs_ea)
        if wt is None:
            wt = _to_float_or_none(spec.wt_kg)
        if wt is None:
            continue
        total_weight += wt
        mg_name = _norm_text(spec.material_group.material_group_name if spec.material_group else None)
        if mg_name in pe_names:
            pe_weight += wt
    pp_weight = total_weight - pe_weight
    if pp_weight < 0:
        pp_weight = 0
    return total_weight, pe_weight, pp_weight


def _get_unit_price(session, model, name_field, value: str):
    row = session.scalar(select(model).where(name_field == value, _active(model)))
    if not row:
        return 0.0
    return _to_float_or_none(row.unit_price) or 0.0


def _build_quotation_snapshot(session, customer: Customer, product: Product, has_lami: bool, extra_rows):
    level = _normalize_customer_level(customer.level) or "C"
    level_factor = CUSTOMER_LEVEL_FACTORS.get(level, 1.0)
    total_weight, pe_weight, pp_weight = _sum_product_spec_weight(session, product.id)
    tb_price = _get_unit_price(session, RawMaterialPrice, RawMaterialPrice.material_name, "TB")
    processing_price = _get_unit_price(session, ProcessingPrice, ProcessingPrice.process_name, "Gia công")
    amount_weight = total_weight * (tb_price + processing_price)
    lami_unit_price = 0.429
    amount_lami = lami_unit_price if has_lami else 0.0
    color_code = (_upper_or_none(product.color) or "").strip()
    amount_color = 0.0 if color_code in {"BG", "WT"} else 0.1
    extras = _sanitize_extra_rows(extra_rows)
    amount_extra = sum(float(r.get("amount") or 0) for r in extras)
    subtotal = amount_weight + amount_lami + amount_color + amount_extra
    total = subtotal * level_factor
    payload = {
        "rows": [
            {"name": "customer_code", "value": customer.customer_code, "amount": None},
            {"name": "product_code", "value": product.product_code, "amount": None},
            {"name": "size", "value": product.spec_inner, "amount": None},
            {"name": "weight", "value": total_weight, "amount": amount_weight},
            {"name": "pe", "value": pe_weight, "amount": None},
            {"name": "pp", "value": pp_weight, "amount": None},
            {"name": "lami", "value": "Y" if has_lami else "N", "amount": amount_lami},
            {"name": "color", "value": product.color, "amount": amount_color},
        ],
        "extra_rows": extras,
    }
    return {
        "has_lami": has_lami,
        "lami_unit_price": lami_unit_price,
        "level_code": level,
        "level_factor": level_factor,
        "size_value": _str_or_none(product.spec_inner),
        "total_weight_kg": total_weight,
        "pe_weight_kg": pe_weight,
        "pp_weight_kg": pp_weight,
        "amount_weight": amount_weight,
        "amount_lami": amount_lami,
        "amount_color": amount_color,
        "amount_extra": amount_extra,
        "subtotal": subtotal,
        "total": total,
        "row_payload": json.dumps(payload, ensure_ascii=False),
    }


def serialize_quotation(item: Quotation):
    payload = {}
    raw_payload = _str_or_none(item.row_payload)
    if raw_payload:
        try:
            payload = json.loads(raw_payload)
        except Exception:
            payload = {}
    return {
        "id": item.id,
        "customer_id": item.customer_id,
        "product_id": item.product_id,
        "has_lami": bool(item.has_lami),
        "lami_unit_price": to_num(item.lami_unit_price),
        "level_code": item.level_code,
        "level_factor": to_num(item.level_factor),
        "size_value": item.size_value,
        "total_weight_kg": to_num(item.total_weight_kg),
        "pe_weight_kg": to_num(item.pe_weight_kg),
        "pp_weight_kg": to_num(item.pp_weight_kg),
        "amount_weight": to_num(item.amount_weight),
        "amount_lami": to_num(item.amount_lami),
        "amount_color": to_num(item.amount_color),
        "amount_extra": to_num(item.amount_extra),
        "subtotal": to_num(item.subtotal),
        "total": to_num(item.total),
        "row_payload": payload,
        "note": item.note,
        "created_at": fmt_datetime(item.created_at),
        "updated_at": fmt_datetime(item.updated_at),
    }


def serialize_spec(s: ProductSpec):
    computed_qty = _compute_qty_from_item_size(s.item_size)
    qty_value = computed_qty if computed_qty is not None else to_num(s.qty_m_or_m2)
    wt_value = _compute_wt_kg(s.unit_weight_kg, qty_value, s.pcs_ea)
    if wt_value is None:
        wt_value = to_num(s.wt_kg)
    return {
        "id": s.id,
        "product_id": s.product_id,
        "item_id": s.item_id,
        "material_group_id": s.material_group_id,
        "line_no": s.line_no,
        "item_name": s.item.item_name if s.item else None,
        "material_group": s.material_group.material_group_name if s.material_group else None,
        "spec": s.spec,
        "item_size": s.item_size,
        "lami": s.lami,
        "item_color": s.item_color,
        "unit_weight_kg": to_num(s.unit_weight_kg),
        "qty_m_or_m2": qty_value,
        "pcs_ea": to_num(s.pcs_ea),
        "wt_kg": wt_value,
        "other_note": s.other_note,
        "is_manual_weight": s.is_manual_weight,
        "created_at": fmt_datetime(s.created_at),
        "updated_at": fmt_datetime(s.updated_at),
    }


def _get_or_create_item(session, item_name: str | None) -> Item | None:
    name = _str_or_none(item_name)
    if not name:
        return None
    item = session.scalar(select(Item).where(Item.item_name == name))
    if item:
        if item.deleted_at:
            item.deleted_at = None
            session.add(item)
        return item
    item = Item(item_name=name)
    session.add(item)
    session.flush()
    return item


def _get_or_create_material_group(session, name: str | None) -> MaterialGroup | None:
    mg_name = _str_or_none(name)
    if not mg_name:
        return None
    mg = session.scalar(select(MaterialGroup).where(MaterialGroup.material_group_name == mg_name))
    if mg:
        if mg.deleted_at:
            mg.deleted_at = None
            session.add(mg)
        return mg
    mg = MaterialGroup(material_group_name=mg_name)
    session.add(mg)
    session.flush()
    return mg


def serialize_plan(p: ProductionPlan):
    return {
        "id": p.id,
        "customer_id": p.customer_id,
        "product_id": p.product_id,
        "lot_no": p.lot_no,
        "etd": fmt_date(p.etd),
        "eta": fmt_date(p.eta),
        "contp_date": fmt_date(p.contp_date),
        "order_qty_pcs": p.order_qty_pcs,
        "spec_inner_snapshot": p.spec_inner_snapshot,
        "liner_snapshot": p.liner_snapshot,
        "print_snapshot": p.print_snapshot,
        "label": p.label,
        "sewing_type": p.sewing_type,
        "packing": p.packing,
        "note": p.note,
        "status": p.status,
        "update_person": p.update_person,
        "created_at": fmt_datetime(p.created_at),
        "updated_at": fmt_datetime(p.updated_at),
    }


def serialize_version(v: ProductPrintVersion):
    return {
        "id": v.id,
        "product_id": v.product_id,
        "version_no": v.version_no,
        "upload_note": v.upload_note,
        "created_by": v.created_by,
        "created_at": fmt_datetime(v.created_at),
        "updated_at": fmt_datetime(v.updated_at),
    }


def serialize_image(i: ProductPrintImage):
    return {
        "id": i.id,
        "product_print_version_id": i.product_print_version_id,
        "image_url": i.image_url,
        "file_name": i.file_name,
        "mime_type": i.mime_type,
        "file_size": i.file_size,
        "width_px": i.width_px,
        "height_px": i.height_px,
        "sort_order": i.sort_order,
        "created_at": fmt_datetime(i.created_at),
        "updated_at": fmt_datetime(i.updated_at),
    }


def serialize_material_group(mg: MaterialGroup):
    effective_unit_weight_value = (
        mg.unit_weight_option.unit_weight_value
        if mg.unit_weight_mode == "choice" and mg.unit_weight_option is not None
        else mg.unit_weight_value
    )
    computed_unit_weight = _compute_unit_weight(
        mg.unit_weight_mode,
        mg.unit_weight_value,
        mg.unit_weight_formula_code,
        mg.spec_label,
        mg.unit_weight_option.unit_weight_value if mg.unit_weight_option is not None else None,
        mg.use_lami_for_calc,
        mg.lami_calc_value,
    )
    return {
        "id": mg.id,
        "material_group_name": mg.material_group_name,
        "spec_label": mg.spec_label,
        "has_lami": mg.has_lami,
        "use_lami_for_calc": mg.use_lami_for_calc,
        "lami_calc_value": to_num(mg.lami_calc_value),
        "pcs_ea_label": mg.pcs_ea_label,
        "unit_weight_mode": mg.unit_weight_mode,
        "unit_weight_value": to_num(effective_unit_weight_value),
        "unit_weight_formula_code": mg.unit_weight_formula_code,
        "unit_weight_formula": mg.unit_weight_formula_code,
        "unit_weight_option_id": mg.unit_weight_option_id,
        "unit_weight_option_label": mg.unit_weight_option.option_label if mg.unit_weight_option else None,
        "unit_weight_option_group": mg.unit_weight_option.option_group if mg.unit_weight_option else None,
        "unit_weight_computed": to_num(computed_unit_weight),
        "unit_weight_note": mg.unit_weight_note,
        "created_at": fmt_datetime(mg.created_at),
        "updated_at": fmt_datetime(mg.updated_at),
    }


def _format_decimal_text(v):
    num = to_num(v)
    if num is None:
        return ""
    text = f"{num:.5f}".rstrip("0").rstrip(".")
    return text


def _apply_form_specification_sheet(ws, product: Product, customer: Customer | None, specs: list[ProductSpec]):
    ws["C2"] = product.product_name or ""
    ws["F2"] = customer.customer_code if customer else ""
    ws["C3"] = product.spec_inner or ""
    ws["F3"] = fmt_date(product.updated_at) or datetime.now().strftime("%d-%m-%Y")
    ws["I3"] = ""
    ws["D4"] = product.top or ""
    ws["F4"] = product.bottom or ""

    # Keep template rows and replace from row 7 onward.
    start_row = 7
    end_row = 21
    for i, spec in enumerate(specs[: end_row - start_row + 1], start=0):
        row_no = start_row + i
        ws.cell(row=row_no, column=1, value=i + 1)
        ws.cell(row=row_no, column=2, value=spec.item.item_name if spec.item else "")
        ws.cell(row=row_no, column=3, value=spec.spec or "")
        ws.cell(row=row_no, column=4, value=spec.item_color or "")
        ws.cell(row=row_no, column=5, value=spec.lami or "")
        ws.cell(row=row_no, column=6, value=spec.item_size or "")
        ws.cell(row=row_no, column=7, value=_format_decimal_text(spec.unit_weight_kg))
        ws.cell(row=row_no, column=8, value=_format_decimal_text(spec.qty_m_or_m2))
        ws.cell(row=row_no, column=9, value=_format_decimal_text(spec.wt_kg))
        ws.cell(row=row_no, column=10, value=spec.other_note or "")

    # Clear remaining template rows in spec table range.
    filled = min(len(specs), end_row - start_row + 1)
    for row_no in range(start_row + filled, end_row + 1):
        for col in range(1, 11):
            ws.cell(row=row_no, column=col, value="")

    total_row = 22
    ws.cell(row=total_row, column=1, value="Tổng")
    total_wt = 0.0
    for spec in specs:
        wt = to_num(spec.wt_kg)
        if wt is None:
            wt = _compute_wt_kg(spec.unit_weight_kg, spec.qty_m_or_m2, spec.pcs_ea)
        if wt is not None:
            total_wt += wt
    ws.cell(row=total_row, column=9, value=_format_decimal_text(total_wt))


def _resolve_media_file_from_url(image_url: str | None):
    url = _str_or_none(image_url)
    if not url:
        return None
    parsed_path = unquote(urlparse(url).path or "")
    media_url = settings.MEDIA_URL or "/media/"
    if not media_url.startswith("/"):
        media_url = f"/{media_url}"
    if not media_url.endswith("/"):
        media_url = f"{media_url}/"

    rel = None
    if parsed_path.startswith(media_url):
        rel = parsed_path[len(media_url) :]
    else:
        marker = "/media/"
        idx = parsed_path.find(marker)
        if idx >= 0:
            rel = parsed_path[idx + len(marker) :]
    if not rel:
        return None
    return Path(settings.MEDIA_ROOT) / rel


def _apply_form_product_sheet(
    ws,
    product: Product,
    customer: Customer | None,
    specs: list[ProductSpec],
    print_images: list[ProductPrintImage],
):
    ws["A1"] = f"Bảng thông tin sản phẩm - {product.product_name or product.product_code or ''}".strip()
    ws["B2"] = product.product_name or ""
    ws["D2"] = product.product_code or ""
    ws["B3"] = customer.customer_code if customer else ""
    ws["D3"] = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    ws["B4"] = product.swl or ""
    ws["D4"] = product.type or ""
    ws["B5"] = product.sewing_type or ""
    ws["D5"] = product.color or ""
    ws["B6"] = product.spec_other or ""
    ws["D6"] = product.spec_inner or ""
    ws["B7"] = product.liner or ""
    ws["D7"] = product.print or ""
    ws["B8"] = product.top or ""
    ws["D8"] = product.bottom or ""
    ws["B9"] = product.packing or ""
    ws["A10"] = "Item"
    ws["B10"] = "Spec"
    ws["C10"] = "Color"
    ws["D10"] = "Item Size"
    ws["A20"] = f"Other : {product.other_note or ''}".strip()

    # In template Form_Product, range E2:J19 is merged for print preview blocks.
    # Only write selected specs into safe columns A/C in row 11..18.
    start_row = 11
    end_row = 18
    for row_no in range(start_row, end_row + 1):
        ws.cell(row=row_no, column=1, value="")
        ws.cell(row=row_no, column=2, value="")
        ws.cell(row=row_no, column=3, value="")
        ws.cell(row=row_no, column=4, value="")

    for i, spec in enumerate(specs[: end_row - start_row + 1], start=0):
        row_no = start_row + i
        ws.cell(row=row_no, column=1, value=spec.item.item_name if spec.item else "")
        ws.cell(row=row_no, column=2, value=spec.spec or "")
        ws.cell(row=row_no, column=3, value=spec.item_color or "")
        ws.cell(row=row_no, column=4, value=spec.item_size or "")

    # Print 1 / Print 2 preview images
    try:
        from openpyxl.drawing.image import Image as XLImage  # type: ignore
    except Exception:
        return

    slots = [("E3", 0), ("H3", 1)]
    for anchor, idx in slots:
        if idx >= len(print_images):
            continue
        abs_path = _resolve_media_file_from_url(print_images[idx].image_url)
        if abs_path is None or not abs_path.exists():
            continue
        try:
            pic = XLImage(str(abs_path))
            pic.width = 210
            pic.height = 300
            ws.add_image(pic, anchor)
        except Exception:
            continue

def serialize_unit_weight_option(item: UnitWeightOption):
    return {
        "id": item.id,
        "option_group": item.option_group,
        "option_label": item.option_label,
        "unit_weight_value": to_num(item.unit_weight_value),
        "created_at": fmt_datetime(item.created_at),
        "updated_at": fmt_datetime(item.updated_at),
    }


def serialize_fixed_weight_table(item: FixedWeightTable):
    return {
        "id": item.id,
        "material_id": item.material_id,
        "material_name": item.material.material_name if item.material else None,
        "material_category_name": (
            item.material.material_category.material_category_name
            if item.material and item.material.material_category
            else None
        ),
        "size_label": item.size_label,
        "unit_weight_value": to_num(item.unit_weight_value),
        "unit_price": to_num(item.unit_price),
        "created_at": fmt_datetime(item.created_at),
        "updated_at": fmt_datetime(item.updated_at),
    }


def serialize_material_category(item: MaterialCategory):
    return {
        "id": item.id,
        "material_category_name": item.material_category_name,
        "material_category_code": item.material_category_code,
        "spec_format": item.spec_format,
        "format": item.format_value,
        "created_at": fmt_datetime(item.created_at),
        "updated_at": fmt_datetime(item.updated_at),
    }


def serialize_material(item: Material):
    return {
        "id": item.id,
        "material_name": item.material_name,
        "material_category_id": item.material_category_id,
        "material_category_name": item.material_category.material_category_name if item.material_category else None,
        "material_category_code": item.material_category.material_category_code if item.material_category else None,
        "formula": item.formula,
        "lami": bool(item.lami),
        "created_at": fmt_datetime(item.created_at),
        "updated_at": fmt_datetime(item.updated_at),
    }


def soft_delete_customer(session, customer_id: int):
    ts = _now()
    product_ids = session.scalars(
        select(Product.id).where(Product.customer_id == customer_id, _active(Product))
    ).all()

    if product_ids:
        version_ids = session.scalars(
            select(ProductPrintVersion.id).where(ProductPrintVersion.product_id.in_(product_ids), _active(ProductPrintVersion))
        ).all()
        if version_ids:
            session.execute(
                update(ProductPrintImage)
                .where(ProductPrintImage.product_print_version_id.in_(version_ids), _active(ProductPrintImage))
                .values(deleted_at=ts)
            )
            session.execute(
                update(ProductPrintVersion)
                .where(ProductPrintVersion.id.in_(version_ids), _active(ProductPrintVersion))
                .values(deleted_at=ts)
            )

        session.execute(
            update(ProductSpec).where(ProductSpec.product_id.in_(product_ids), _active(ProductSpec)).values(deleted_at=ts)
        )
        session.execute(update(Product).where(Product.id.in_(product_ids), _active(Product)).values(deleted_at=ts))
        session.execute(
            update(ProductionPlan)
            .where(ProductionPlan.product_id.in_(product_ids), _active(ProductionPlan))
            .values(deleted_at=ts)
        )

    session.execute(
        update(ProductionPlan)
        .where(ProductionPlan.customer_id == customer_id, _active(ProductionPlan))
        .values(deleted_at=ts)
    )
    session.execute(update(Customer).where(Customer.id == customer_id).values(deleted_at=ts))


def soft_delete_product(session, product_id: int):
    ts = _now()
    version_ids = session.scalars(
        select(ProductPrintVersion.id).where(ProductPrintVersion.product_id == product_id, _active(ProductPrintVersion))
    ).all()
    if version_ids:
        session.execute(
            update(ProductPrintImage)
            .where(ProductPrintImage.product_print_version_id.in_(version_ids), _active(ProductPrintImage))
            .values(deleted_at=ts)
        )
        session.execute(
            update(ProductPrintVersion)
            .where(ProductPrintVersion.id.in_(version_ids), _active(ProductPrintVersion))
            .values(deleted_at=ts)
        )

    session.execute(update(ProductSpec).where(ProductSpec.product_id == product_id, _active(ProductSpec)).values(deleted_at=ts))
    session.execute(
        update(ProductionPlan).where(ProductionPlan.product_id == product_id, _active(ProductionPlan)).values(deleted_at=ts)
    )
    session.execute(update(Product).where(Product.id == product_id).values(deleted_at=ts))


def soft_delete_print_version(session, version_id: int):
    ts = _now()
    session.execute(
        update(ProductPrintImage)
        .where(ProductPrintImage.product_print_version_id == version_id, _active(ProductPrintImage))
        .values(deleted_at=ts)
    )
    session.execute(update(ProductPrintVersion).where(ProductPrintVersion.id == version_id).values(deleted_at=ts))


def _sync_product_has_print_assets(session, product_id: int):
    active_images = session.scalar(
        select(func.count(ProductPrintImage.id))
        .join(ProductPrintVersion, ProductPrintImage.product_print_version_id == ProductPrintVersion.id)
        .where(
            ProductPrintVersion.product_id == product_id,
            _active(ProductPrintVersion),
            _active(ProductPrintImage),
        )
    ) or 0
    session.execute(update(Product).where(Product.id == product_id).values(has_print_assets=active_images > 0))


@csrf_exempt
def health(_request: HttpRequest):
    _ensure_material_master_tables()
    return _ok({"status": "ok"})


@csrf_exempt
def login(request: HttpRequest):
    _ensure_material_master_tables()
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)
    body = _body(request)
    username = (body.get("username") or "").strip()
    password = (body.get("password") or "").strip()
    result = login_with_username_password(username, password)
    if not result:
        return _ok({"detail": "Tên đăng nhập hoặc mật khẩu không đúng"}, 401)
    token, user = result
    return _ok({"token": token, "user": user})


@csrf_exempt
@require_auth
def me(request: HttpRequest):
    if request.method != "GET":
        return _ok({"detail": "Method not allowed"}, 405)
    u = request.current_user
    return _ok(serialize_user(u))


@csrf_exempt
@require_auth
def me_update(request: HttpRequest):
    if request.method != "PUT":
        return _ok({"detail": "Method not allowed"}, 405)
    body = _body(request)
    with get_session() as session:
        item = session.scalar(select(User).where(User.id == request.current_user.id, _active(User)))
        if not item:
            return _ok({"detail": "Not found"}, 404)
        if "full_name" in body:
            item.full_name = body.get("full_name")
        if "avatar_url" in body:
            item.avatar_url = body.get("avatar_url")
        if "role" in body:
            if not _is_admin(request.current_user):
                return _forbidden()
            if body.get("role") not in {"admin", "manager", "staff"}:
                return _ok({"detail": "Role không hợp lệ"}, 400)
            item.role = body.get("role")
        session.add(item)
        session.flush()
        return _ok(serialize_user(item))


@csrf_exempt
@require_auth
def change_password(request: HttpRequest):
    if request.method != "PUT":
        return _ok({"detail": "Method not allowed"}, 405)
    body = _body(request)
    current_password = (body.get("current_password") or "").strip()
    new_password = (body.get("new_password") or "").strip()
    if len(new_password) < 6:
        return _ok({"detail": "Mật khẩu mới tối thiểu 6 ký tự"}, 400)
    with get_session() as session:
        item = session.scalar(select(User).where(User.id == request.current_user.id, _active(User)))
        if not item:
            return _ok({"detail": "Not found"}, 404)
        if not check_password(current_password, item.password_hash):
            return _ok({"detail": "Mật khẩu hiện tại không đúng"}, 400)
        item.password_hash = make_password(new_password)
        session.add(item)
        session.flush()
        return _ok({"success": True})


@csrf_exempt
@require_auth
def logout(request: HttpRequest):
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)
    auth = request.headers.get("Authorization", "")
    token_value = auth.replace("Token ", "", 1).strip() if auth.startswith("Token ") else ""
    if not token_value:
        return _ok({"success": True})
    with get_session() as session:
        token = session.scalar(select(AuthToken).where(AuthToken.token == token_value))
        if token:
            session.delete(token)
    return _ok({"success": True})


@csrf_exempt
@require_auth
def users(request: HttpRequest):
    if not _is_admin(request.current_user):
        return _forbidden()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip()
            q = select(User).where(_active(User))
            if search:
                like = f"%{search}%"
                q = q.where((User.username.ilike(like)) | (User.full_name.ilike(like)))
            rows = session.scalars(q.order_by(User.id.desc())).all()
            return _ok([serialize_user(r) for r in rows])
        if request.method == "POST":
            body = _body(request)
            username = (body.get("username") or "").strip()
            password = (body.get("password") or "").strip()
            role = (body.get("role") or "staff").strip()
            if not username or not password:
                return _ok({"detail": "Thiếu username/password"}, 400)
            if role not in {"admin", "manager", "staff"}:
                return _ok({"detail": "Role không hợp lệ"}, 400)
            exists = session.scalar(select(User).where(User.username == username, _active(User)))
            if exists:
                return _ok({"detail": "Username đã tồn tại"}, 400)
            item = User(
                username=username,
                password_hash=make_password(password),
                full_name=body.get("full_name"),
                avatar_url=body.get("avatar_url"),
                role=role,
                is_active=bool(body.get("is_active", True)),
            )
            session.add(item)
            session.flush()
            return _ok(serialize_user(item), 201)
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def user_detail(request: HttpRequest, user_id: int):
    if not _is_admin(request.current_user):
        return _forbidden()
    with get_session() as session:
        item = session.scalar(select(User).where(User.id == user_id, _active(User)))
        if not item:
            return _ok({"detail": "Not found"}, 404)
        if request.method == "GET":
            return _ok(serialize_user(item))
        if request.method == "PUT":
            body = _body(request)
            if "username" in body:
                username = (body.get("username") or "").strip()
                if not username:
                    return _ok({"detail": "Username không hợp lệ"}, 400)
                dup = session.scalar(select(User).where(User.username == username, User.id != item.id, _active(User)))
                if dup:
                    return _ok({"detail": "Username đã tồn tại"}, 400)
                item.username = username
            if "password" in body and body.get("password"):
                item.password_hash = make_password(body.get("password"))
            if "full_name" in body:
                item.full_name = body.get("full_name")
            if "avatar_url" in body:
                item.avatar_url = body.get("avatar_url")
            if "role" in body:
                role = body.get("role")
                if role not in {"admin", "manager", "staff"}:
                    return _ok({"detail": "Role không hợp lệ"}, 400)
                item.role = role
            if "is_active" in body:
                item.is_active = bool(body.get("is_active"))
            session.add(item)
            session.flush()
            return _ok(serialize_user(item))
        if request.method == "DELETE":
            if item.username == "admin":
                return _ok({"detail": "Không thể xóa tài khoản admin mặc định"}, 400)
            item.deleted_at = _now()
            item.is_active = False
            session.add(item)
            tokens = session.scalars(select(AuthToken).where(AuthToken.user_id == item.id)).all()
            for tk in tokens:
                session.delete(tk)
            session.flush()
            return _ok({"success": True})
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def product_types(request: HttpRequest):
    _ensure_product_types_table()
    with get_session() as session:
        if request.method == "GET":
            rows = session.scalars(
                select(ProductType)
                .where(_active(ProductType), ProductType.product_type_name != "OTHER")
                .order_by(ProductType.id.asc())
            ).all()
            return _ok([serialize_product_type(r) for r in rows])
        if request.method == "POST":
            body = _body(request)
            name = _str_or_none(body.get("product_type_name"))
            if not name:
                return _ok({"detail": "Thiếu product_type_name"}, 400)
            normalized = name.upper()
            if normalized == "OTHER":
                return _ok({"detail": "Không hỗ trợ loại sản phẩm OTHER"}, 400)
            existing = session.scalar(select(ProductType).where(ProductType.product_type_name == normalized))
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Loại sản phẩm đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.product_type_name = normalized
                session.add(existing)
                session.flush()
                return _ok(serialize_product_type(existing))
            row = ProductType(product_type_name=normalized)
            session.add(row)
            session.flush()
            return _ok(serialize_product_type(row), 201)
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def product_type_detail(request: HttpRequest, item_id: int):
    _ensure_product_types_table()
    with get_session() as session:
        row = session.scalar(select(ProductType).where(ProductType.id == item_id, _active(ProductType)))
        if not row:
            return _ok({"detail": "Not found"}, 404)
        if request.method == "PUT":
            body = _body(request)
            name = _str_or_none(body.get("product_type_name"))
            if not name:
                return _ok({"detail": "Thiếu product_type_name"}, 400)
            normalized = name.upper()
            if normalized == "OTHER":
                return _ok({"detail": "Không hỗ trợ loại sản phẩm OTHER"}, 400)
            dup = session.scalar(
                select(ProductType).where(
                    ProductType.product_type_name == normalized,
                    ProductType.id != row.id,
                    _active(ProductType),
                )
            )
            if dup:
                return _ok({"detail": "Loại sản phẩm đã tồn tại"}, 400)
            row.product_type_name = normalized
            session.add(row)
            session.flush()
            return _ok(serialize_product_type(row))
        if request.method == "DELETE":
            in_use = session.scalar(
                select(func.count()).select_from(Product).where(Product.type == row.product_type_name, _active(Product))
            )
            if in_use and int(in_use) > 0:
                return _ok({"detail": "Loại sản phẩm đang được sử dụng trong Products"}, 400)
            row.deleted_at = _now()
            session.add(row)
            session.flush()
            return _ok({"success": True})
    return _ok({"detail": "Method not allowed"}, 405)


def serialize_item_type_formula(row: ItemTypeFormula):
    return {
        "item_id": row.item_id,
        "product_type_id": row.product_type_id,
        "formula": row.formula,
        "created_at": fmt_datetime(row.created_at),
        "updated_at": fmt_datetime(row.updated_at),
    }


@csrf_exempt
@require_auth
def item_type_formulas(request: HttpRequest):
    _ensure_items_table_schema()
    _ensure_product_types_table()
    with get_session() as session:
        if request.method == "GET":
            configured_ids = session.scalars(select(ItemTypeFormulaItem.item_id)).all()
            configured_id_set = {int(x) for x in configured_ids}
            all_items = session.scalars(select(Item).where(_active(Item)).order_by(Item.item_name.asc())).all()
            items = [x for x in all_items if x.id in configured_id_set]
            available_items = [x for x in all_items if x.id not in configured_id_set]
            product_types = session.scalars(
                select(ProductType)
                .where(_active(ProductType), ProductType.product_type_name != "OTHER")
                .order_by(ProductType.id.asc())
            ).all()
            formulas = session.scalars(
                select(ItemTypeFormula)
                .join(Item, Item.id == ItemTypeFormula.item_id)
                .join(ProductType, ProductType.id == ItemTypeFormula.product_type_id)
                .where(
                    _active(Item),
                    _active(ProductType),
                    ProductType.product_type_name != "OTHER",
                    ItemTypeFormula.item_id.in_(configured_id_set if configured_id_set else {0}),
                )
            ).all()
            return _ok(
                {
                    "items": [{"id": x.id, "item_name": x.item_name} for x in items],
                    "available_items": [{"id": x.id, "item_name": x.item_name} for x in available_items],
                    "product_types": [serialize_product_type(x) for x in product_types],
                    "formulas": [serialize_item_type_formula(x) for x in formulas],
                }
            )
        if request.method == "POST":
            body = _body(request)
            raw_ids = body.get("item_ids")
            if not isinstance(raw_ids, list) or len(raw_ids) == 0:
                return _ok({"detail": "Thiếu item_ids"}, 400)
            created = 0
            for raw_id in raw_ids:
                try:
                    item_id = int(raw_id)
                except (TypeError, ValueError):
                    continue
                item = session.scalar(select(Item).where(Item.id == item_id, _active(Item)))
                if not item:
                    continue
                exists = session.scalar(
                    select(ItemTypeFormulaItem).where(ItemTypeFormulaItem.item_id == item_id)
                )
                if exists:
                    continue
                session.add(ItemTypeFormulaItem(item_id=item_id))
                created += 1
            session.flush()
            return _ok({"success": True, "created": created})
        if request.method == "PUT":
            body = _body(request)
            item_id = body.get("item_id")
            product_type_id = body.get("product_type_id")
            formula = _str_or_none(body.get("formula"))
            if not item_id or not product_type_id:
                return _ok({"detail": "Thiếu item_id hoặc product_type_id"}, 400)
            item = session.scalar(select(Item).where(Item.id == int(item_id), _active(Item)))
            if not item:
                return _ok({"detail": "item_id không hợp lệ"}, 400)
            pt = session.scalar(select(ProductType).where(ProductType.id == int(product_type_id), _active(ProductType)))
            if not pt:
                return _ok({"detail": "product_type_id không hợp lệ"}, 400)
            existing = session.scalar(
                select(ItemTypeFormula).where(
                    ItemTypeFormula.item_id == item.id,
                    ItemTypeFormula.product_type_id == pt.id,
                )
            )
            if not formula:
                if existing:
                    session.delete(existing)
                    session.flush()
                return _ok({"success": True})
            pair = _split_pair_formula(formula)
            if not pair:
                return _ok({"detail": "Công thức không hợp lệ. Định dạng đúng: left x right"}, 400)
            left_expr, right_expr = pair
            left_valid = _validate_formula_expr(left_expr)
            right_valid = _validate_formula_expr(right_expr)
            if not left_valid or not right_valid:
                return _ok({"detail": "Công thức không hợp lệ. Chỉ dùng A/B/C, số và + - * / ( )"}, 400)
            validated = f"{left_valid} x {right_valid}"
            if existing:
                existing.formula = validated
                session.add(existing)
                session.flush()
                return _ok(serialize_item_type_formula(existing))
            row = ItemTypeFormula(item_id=item.id, product_type_id=pt.id, formula=validated)
            session.add(row)
            session.flush()
            return _ok(serialize_item_type_formula(row), 201)
        if request.method == "DELETE":
            body = _body(request)
            item_id = body.get("item_id")
            if not item_id:
                return _ok({"detail": "Thiếu item_id"}, 400)
            try:
                target_item_id = int(item_id)
            except (TypeError, ValueError):
                return _ok({"detail": "item_id không hợp lệ"}, 400)
            config = session.scalar(select(ItemTypeFormulaItem).where(ItemTypeFormulaItem.item_id == target_item_id))
            if config:
                session.delete(config)
            formula_rows = session.scalars(select(ItemTypeFormula).where(ItemTypeFormula.item_id == target_item_id)).all()
            for row in formula_rows:
                session.delete(row)
            session.flush()
            return _ok({"success": True})
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def customers(request: HttpRequest):
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip()
            q = select(Customer).where(_active(Customer))
            if search:
                like = f"%{search}%"
                q = q.where(
                    (Customer.customer_code.ilike(like))
                    | (Customer.customer_name.ilike(like))
                    | (Customer.phone.ilike(like))
                    | (Customer.email.ilike(like))
                )
            rows = session.scalars(q.order_by(Customer.id.desc())).all()
            return _ok([serialize_customer(r) for r in rows])

        if request.method == "POST":
            body = _body(request)
            level_input = body.get("level")
            level = _normalize_customer_level(level_input)
            if _str_or_none(level_input) and level is None:
                return _ok({"detail": "Cấp độ khách hàng không hợp lệ. Chỉ chấp nhận A/B/C/N"}, 400)
            item = Customer(
                customer_code=body["customer_code"],
                customer_name=body["customer_name"],
                address=body.get("address"),
                contact_person=body.get("contact_person"),
                phone=body.get("phone"),
                email=body.get("email"),
                production_2025=body.get("production_2025") or 0,
                production_2026=body.get("production_2026") or 0,
                in_production=body.get("in_production") or 0,
                level=level,
            )
            session.add(item)
            session.flush()
            return _ok(serialize_customer(item), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def customers_import_excel(request: HttpRequest):
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)

    excel_file = request.FILES.get("file")
    if not excel_file:
        return _ok({"detail": "Thiếu file Excel"}, 400)
    if not excel_file.name.lower().endswith(".xlsx"):
        return _ok({"detail": "Chỉ hỗ trợ file .xlsx"}, 400)

    tmp_dir = Path(settings.MEDIA_ROOT) / "tmp_import"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"{uuid.uuid4().hex}.xlsx"
    with tmp_path.open("wb+") as dst:
        for chunk in excel_file.chunks():
            dst.write(chunk)

    try:
        rows = _xlsx_sheet_rows(str(tmp_path), "Customers")
    except ValueError as ex:
        tmp_path.unlink(missing_ok=True)
        return _ok({"detail": str(ex)}, 400)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        return _ok({"detail": "Không đọc được file Excel"}, 400)
    finally:
        tmp_path.unlink(missing_ok=True)

    if not rows:
        return _ok({"detail": "Sheet Customers không có dữ liệu"}, 400)

    headers = [(_str_or_none(v) or "") for v in rows[0]]
    idx = {h.strip().lower(): i for i, h in enumerate(headers) if h}
    required = {"customercode", "customername"}
    if not required.issubset(set(idx.keys())):
        return _ok({"detail": "Thiếu cột bắt buộc CustomerCode/CustomerName"}, 400)

    def val(row_values: list[str | None], key: str):
        pos = idx.get(key)
        if pos is None or pos >= len(row_values):
            return None
        return row_values[pos]

    created = 0
    skipped = 0
    failed: list[dict] = []

    with get_session() as session:
        existing_rows = session.execute(select(Customer.customer_code, Customer.email, Customer.phone, Customer.deleted_at)).all()
        existing_codes = {_norm_text(r[0]) for r in existing_rows if r[0]}
        existing_emails = {_norm_text(r[1]) for r in existing_rows if r[1] and r[3] is None}
        existing_phones = {_norm_phone(r[2]) for r in existing_rows if r[2] and r[3] is None}

        seen_codes: set[str] = set()
        seen_emails: set[str] = set()
        seen_phones: set[str] = set()

        for line_no, row in enumerate(rows[1:], start=2):
            code = _str_or_none(val(row, "customercode"))
            name = _str_or_none(val(row, "customername"))
            email = _str_or_none(val(row, "email"))
            phone = _str_or_none(val(row, "phone"))

            if not code and not name:
                skipped += 1
                continue
            reasons: list[str] = []
            if not code:
                reasons.append("Thiếu mã khách hàng")
            if not name:
                reasons.append("Thiếu tên khách hàng")

            code_key = _norm_text(code)
            email_key = _norm_text(email)
            phone_key = _norm_phone(phone)

            if code_key and (code_key in existing_codes or code_key in seen_codes):
                reasons.append("Trùng mã khách hàng")
            if email_key and (email_key in existing_emails or email_key in seen_emails):
                reasons.append("Trùng email")
            if phone_key and (phone_key in existing_phones or phone_key in seen_phones):
                reasons.append("Trùng số điện thoại")
            level_raw = _str_or_none(val(row, "level"))
            level = _normalize_customer_level(level_raw)
            if level_raw and level is None:
                reasons.append("Cấp độ không hợp lệ (chỉ A/B/C/N)")

            if reasons:
                failed.append(
                    {
                        "row": line_no,
                        "customer_code": code,
                        "customer_name": name,
                        "reasons": reasons,
                    }
                )
                continue

            payload = {
                "customer_code": code,
                "customer_name": name,
                "address": _str_or_none(val(row, "address")),
                "contact_person": _str_or_none(val(row, "contactperson")),
                "phone": phone,
                "email": email,
                "production_2025": _num_or_zero(val(row, "production_2025")),
                "production_2026": _num_or_zero(val(row, "production_2026")),
                "in_production": _num_or_zero(val(row, "in_production")),
                "level": level,
            }
            session.add(Customer(**payload))
            created += 1
            seen_codes.add(code_key)
            if email_key:
                seen_emails.add(email_key)
            if phone_key:
                seen_phones.add(phone_key)

        session.flush()

    return _ok(
        {
            "success": True,
            "created": created,
            "skipped": skipped,
            "failed_count": len(failed),
            "failed": failed[:200],
        }
    )


@csrf_exempt
@require_auth
def customer_detail(request: HttpRequest, item_id: int):
    with get_session() as session:
        item = session.scalar(select(Customer).where(Customer.id == item_id, _active(Customer)))
        if not item:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "GET":
            return _ok(serialize_customer(item))

        if request.method == "PUT":
            body = _body(request)
            if "level" in body:
                level_input = body.get("level")
                level = _normalize_customer_level(level_input)
                if _str_or_none(level_input) and level is None:
                    return _ok({"detail": "Cấp độ khách hàng không hợp lệ. Chỉ chấp nhận A/B/C/N"}, 400)
                body["level"] = level
            for f in [
                "customer_code",
                "customer_name",
                "address",
                "contact_person",
                "phone",
                "email",
                "production_2025",
                "production_2026",
                "in_production",
                "level",
            ]:
                if f in body:
                    setattr(item, f, body[f])
            session.add(item)
            session.flush()
            return _ok(serialize_customer(item))

        if request.method == "DELETE":
            soft_delete_customer(session, item.id)
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def unit_weight_options(request: HttpRequest):
    with get_session() as session:
        if request.method == "GET":
            rows = session.scalars(
                select(UnitWeightOption)
                .where(_active(UnitWeightOption))
                .order_by(UnitWeightOption.option_group.asc(), UnitWeightOption.option_label.asc())
            ).all()
            return _ok([serialize_unit_weight_option(r) for r in rows])

        if request.method == "POST":
            body = _body(request)
            option_group = _str_or_none(body.get("option_group"))
            option_label = _str_or_none(body.get("option_label"))
            value_text = _normalize_number_text(body.get("unit_weight_value"))
            if not option_group:
                return _ok({"detail": "Thiếu option_group"}, 400)
            if not option_label:
                return _ok({"detail": "Thiếu option_label"}, 400)
            if value_text is None:
                return _ok({"detail": "unit_weight_value phải là số"}, 400)

            existing = session.scalar(
                select(UnitWeightOption).where(
                    UnitWeightOption.option_group == option_group,
                    UnitWeightOption.option_label == option_label,
                )
            )
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Option đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.unit_weight_value = value_text
                session.add(existing)
                session.flush()
                return _ok(serialize_unit_weight_option(existing))

            row = UnitWeightOption(
                option_group=option_group,
                option_label=option_label,
                unit_weight_value=value_text,
            )
            session.add(row)
            session.flush()
            return _ok(serialize_unit_weight_option(row), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def unit_weight_option_detail(request: HttpRequest, item_id: int):
    with get_session() as session:
        row = session.scalar(select(UnitWeightOption).where(UnitWeightOption.id == item_id, _active(UnitWeightOption)))
        if not row:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "PUT":
            body = _body(request)
            option_group = _str_or_none(body.get("option_group"))
            option_label = _str_or_none(body.get("option_label"))
            value_text = _normalize_number_text(body.get("unit_weight_value"))
            if not option_group:
                return _ok({"detail": "Thiếu option_group"}, 400)
            if not option_label:
                return _ok({"detail": "Thiếu option_label"}, 400)
            if value_text is None:
                return _ok({"detail": "unit_weight_value phải là số"}, 400)

            dup = session.scalar(
                select(UnitWeightOption).where(
                    UnitWeightOption.option_group == option_group,
                    UnitWeightOption.option_label == option_label,
                    UnitWeightOption.id != row.id,
                    _active(UnitWeightOption),
                )
            )
            if dup:
                return _ok({"detail": "Option đã tồn tại"}, 400)

            row.option_group = option_group
            row.option_label = option_label
            row.unit_weight_value = value_text
            session.add(row)
            session.flush()
            return _ok(serialize_unit_weight_option(row))

        if request.method == "DELETE":
            in_use = session.scalar(
                select(func.count())
                .select_from(MaterialGroup)
                .where(MaterialGroup.unit_weight_option_id == row.id, _active(MaterialGroup))
            )
            if in_use and int(in_use) > 0:
                return _ok({"detail": "Option đang được sử dụng trong Material Group"}, 400)
            row.deleted_at = _now()
            session.add(row)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def fixed_weight_tables(request: HttpRequest):
    _ensure_fixed_weight_tables_table()
    _ensure_material_master_tables()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip().lower()
            material_id = request.GET.get("material_id")
            q = select(FixedWeightTable).where(_active(FixedWeightTable))
            if material_id:
                q = q.where(FixedWeightTable.material_id == int(material_id))
            rows = session.scalars(q.order_by(FixedWeightTable.id.desc())).all()
            out: list[dict] = []
            for row in rows:
                data = serialize_fixed_weight_table(row)
                if search:
                    hay = " ".join(
                        [
                            (data.get("material_name") or ""),
                            (data.get("material_category_name") or ""),
                            (data.get("size_label") or ""),
                        ]
                    ).lower()
                    if search not in hay:
                        continue
                out.append(data)
            return _ok(out)

        if request.method == "POST":
            body = _body(request)
            material_id = body.get("material_id")
            size_label = _str_or_none(body.get("size_label"))
            unit_weight_value = _normalize_number_text(body.get("unit_weight_value"))
            unit_price = _normalize_number_text(body.get("unit_price"))
            if not material_id:
                return _ok({"detail": "Thiếu material_id"}, 400)
            if not size_label:
                return _ok({"detail": "Thiếu size_label"}, 400)
            if unit_weight_value is None:
                return _ok({"detail": "unit_weight_value phải là số"}, 400)
            if unit_price is None:
                return _ok({"detail": "unit_price phải là số"}, 400)
            material = session.scalar(select(Material).where(Material.id == int(material_id), _active(Material)))
            if not material:
                return _ok({"detail": "material_id không hợp lệ"}, 400)

            existing = session.scalar(
                select(FixedWeightTable).where(
                    FixedWeightTable.material_id == material.id,
                    FixedWeightTable.size_label == size_label,
                )
            )
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Dòng định lượng đã tồn tại trong material"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.unit_weight_value = unit_weight_value
                existing.unit_price = unit_price
                session.add(existing)
                session.flush()
                return _ok(serialize_fixed_weight_table(existing))

            row = FixedWeightTable(
                material_id=material.id,
                size_label=size_label,
                unit_weight_value=unit_weight_value,
                unit_price=unit_price,
            )
            session.add(row)
            session.flush()
            return _ok(serialize_fixed_weight_table(row), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def fixed_weight_table_detail(request: HttpRequest, item_id: int):
    _ensure_fixed_weight_tables_table()
    _ensure_material_master_tables()
    with get_session() as session:
        row = session.scalar(select(FixedWeightTable).where(FixedWeightTable.id == item_id, _active(FixedWeightTable)))
        if not row:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "PUT":
            body = _body(request)
            material_id = body.get("material_id")
            size_label = _str_or_none(body.get("size_label"))
            unit_weight_value = _normalize_number_text(body.get("unit_weight_value"))
            unit_price = _normalize_number_text(body.get("unit_price"))
            if not material_id:
                return _ok({"detail": "Thiếu material_id"}, 400)
            if not size_label:
                return _ok({"detail": "Thiếu size_label"}, 400)
            if unit_weight_value is None:
                return _ok({"detail": "unit_weight_value phải là số"}, 400)
            if unit_price is None:
                return _ok({"detail": "unit_price phải là số"}, 400)
            material = session.scalar(select(Material).where(Material.id == int(material_id), _active(Material)))
            if not material:
                return _ok({"detail": "material_id không hợp lệ"}, 400)
            dup = session.scalar(
                select(FixedWeightTable).where(
                    FixedWeightTable.material_id == material.id,
                    FixedWeightTable.size_label == size_label,
                    FixedWeightTable.id != row.id,
                    _active(FixedWeightTable),
                )
            )
            if dup:
                return _ok({"detail": "Dòng định lượng đã tồn tại trong material"}, 400)
            row.material_id = material.id
            row.size_label = size_label
            row.unit_weight_value = unit_weight_value
            row.unit_price = unit_price
            session.add(row)
            session.flush()
            return _ok(serialize_fixed_weight_table(row))

        if request.method == "DELETE":
            row.deleted_at = _now()
            session.add(row)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def material_categories(request: HttpRequest):
    _ensure_material_master_tables()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip().lower()
            rows = session.scalars(select(MaterialCategory).where(_active(MaterialCategory)).order_by(MaterialCategory.material_category_name.asc())).all()
            if search:
                rows = [
                    r
                    for r in rows
                    if search in (r.material_category_name or "").lower()
                    or search in (r.material_category_code or "").lower()
                    or search in (r.spec_format or "").lower()
                    or search in (r.format_value or "").lower()
                ]
            return _ok([serialize_material_category(r) for r in rows])
        if request.method == "POST":
            body = _body(request)
            name = _str_or_none(body.get("material_category_name"))
            category_code = _normalize_material_category_code(_str_or_none(body.get("material_category_code")), name)
            spec_format = (_str_or_none(body.get("spec_format")) or "text").lower()
            format_value = _str_or_none(body.get("format"))
            if not name:
                return _ok({"detail": "Thiếu material_category_name"}, 400)
            if spec_format not in {"size", "text"}:
                return _ok({"detail": "spec_format chỉ chấp nhận size hoặc text"}, 400)
            if spec_format == "size" and not format_value:
                return _ok({"detail": "Thiếu format khi spec_format = size"}, 400)
            if spec_format != "size":
                format_value = None
            existing = session.scalar(select(MaterialCategory).where(MaterialCategory.material_category_name == name))
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Material category đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.spec_format = spec_format
                existing.format_value = format_value
                existing.material_category_code = category_code
                session.add(existing)
                session.flush()
                return _ok(serialize_material_category(existing))
            row = MaterialCategory(
                material_category_name=name,
                material_category_code=category_code,
                spec_format=spec_format,
                format_value=format_value,
            )
            session.add(row)
            session.flush()
            return _ok(serialize_material_category(row), 201)
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def materials(request: HttpRequest):
    _ensure_material_master_tables()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip().lower()
            rows = session.scalars(select(Material).where(_active(Material)).order_by(Material.material_name.asc())).all()
            if search:
                rows = [
                    r
                    for r in rows
                    if search in (r.material_name or "").lower()
                    or search in (r.formula or "").lower()
                    or search in ((r.material_category.material_category_name if r.material_category else "") or "").lower()
                ]
            return _ok([serialize_material(r) for r in rows])
        if request.method == "POST":
            body = _body(request)
            name = _str_or_none(body.get("material_name"))
            material_category_id = body.get("material_category_id")
            formula = _str_or_none(body.get("formula"))
            lami = bool(body.get("lami"))
            if not name:
                return _ok({"detail": "Thiếu material_name"}, 400)
            if not material_category_id:
                return _ok({"detail": "Thiếu material_category_id"}, 400)
            category = session.scalar(
                select(MaterialCategory).where(MaterialCategory.id == int(material_category_id), _active(MaterialCategory))
            )
            if not category:
                return _ok({"detail": "material_category_id không hợp lệ"}, 400)
            if not _is_fabric_category(category.material_category_code, category.material_category_name):
                formula = None
            existing = session.scalar(select(Material).where(Material.material_name == name))
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Material đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.material_category_id = category.id
                existing.formula = formula
                existing.lami = lami
                session.add(existing)
                session.flush()
                return _ok(serialize_material(existing))
            row = Material(material_name=name, material_category_id=category.id, formula=formula, lami=lami)
            session.add(row)
            session.flush()
            return _ok(serialize_material(row), 201)
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def material_category_detail(request: HttpRequest, item_id: int):
    _ensure_material_master_tables()
    with get_session() as session:
        row = session.scalar(select(MaterialCategory).where(MaterialCategory.id == item_id, _active(MaterialCategory)))
        if not row:
            return _ok({"detail": "Not found"}, 404)
        if request.method == "PUT":
            body = _body(request)
            name = _str_or_none(body.get("material_category_name"))
            category_code = _normalize_material_category_code(_str_or_none(body.get("material_category_code")), name)
            spec_format = (_str_or_none(body.get("spec_format")) or "text").lower()
            format_value = _str_or_none(body.get("format"))
            if not name:
                return _ok({"detail": "Thiếu material_category_name"}, 400)
            if spec_format not in {"size", "text"}:
                return _ok({"detail": "spec_format chỉ chấp nhận size hoặc text"}, 400)
            if spec_format == "size" and not format_value:
                return _ok({"detail": "Thiếu format khi spec_format = size"}, 400)
            if spec_format != "size":
                format_value = None
            dup = session.scalar(
                select(MaterialCategory).where(
                    MaterialCategory.material_category_name == name,
                    MaterialCategory.id != row.id,
                    _active(MaterialCategory),
                )
            )
            if dup:
                return _ok({"detail": "Material category đã tồn tại"}, 400)
            row.material_category_name = name
            row.material_category_code = category_code
            row.spec_format = spec_format
            row.format_value = format_value
            session.add(row)
            session.flush()
            return _ok(serialize_material_category(row))
        if request.method == "DELETE":
            row.deleted_at = _now()
            session.add(row)
            session.flush()
            return _ok({"success": True})
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def material_detail(request: HttpRequest, item_id: int):
    _ensure_material_master_tables()
    with get_session() as session:
        row = session.scalar(select(Material).where(Material.id == item_id, _active(Material)))
        if not row:
            return _ok({"detail": "Not found"}, 404)
        if request.method == "PUT":
            body = _body(request)
            name = _str_or_none(body.get("material_name"))
            material_category_id = body.get("material_category_id")
            formula = _str_or_none(body.get("formula"))
            lami = bool(body.get("lami"))
            if not name:
                return _ok({"detail": "Thiếu material_name"}, 400)
            if not material_category_id:
                return _ok({"detail": "Thiếu material_category_id"}, 400)
            category = session.scalar(
                select(MaterialCategory).where(MaterialCategory.id == int(material_category_id), _active(MaterialCategory))
            )
            if not category:
                return _ok({"detail": "material_category_id không hợp lệ"}, 400)
            if not _is_fabric_category(category.material_category_code, category.material_category_name):
                formula = None
            dup = session.scalar(
                select(Material).where(
                    Material.material_name == name,
                    Material.id != row.id,
                    _active(Material),
                )
            )
            if dup:
                return _ok({"detail": "Material đã tồn tại"}, 400)
            row.material_name = name
            row.material_category_id = category.id
            row.formula = formula
            row.lami = lami
            session.add(row)
            session.flush()
            return _ok(serialize_material(row))
        if request.method == "DELETE":
            row.deleted_at = _now()
            session.add(row)
            session.flush()
            return _ok({"success": True})
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def material_groups(request: HttpRequest):
    with get_session() as session:
        if request.method == "GET":
            rows = session.scalars(
                select(MaterialGroup).where(_active(MaterialGroup)).order_by(MaterialGroup.material_group_name.asc())
            ).all()
            return _ok([serialize_material_group(r) for r in rows])

        if request.method == "POST":
            body = _body(request)
            name = _str_or_none(body.get("material_group_name"))
            if not name:
                return _ok({"detail": "Thiếu tên material group"}, 400)
            raw_spec = _str_or_none(body.get("spec_label"))
            raw_pcs = _str_or_none(body.get("pcs_ea_label"))
            spec_label = _normalize_spec_abc(raw_spec)
            pcs_ea_label = _normalize_number_text(raw_pcs) if raw_pcs is not None else "1"
            if raw_spec and not spec_label:
                return _ok({"detail": "Spec phải đúng định dạng A*B*C (A là text hoặc số, B và C là số)"}, 400)
            if raw_pcs and pcs_ea_label is None:
                return _ok({"detail": "PCS (EA) phải là số"}, 400)
            has_lami = bool(body.get("has_lami"))
            use_lami_for_calc = bool(body.get("use_lami_for_calc")) and has_lami
            lami_calc_value = _str_or_none(body.get("lami_calc_value"))
            if use_lami_for_calc:
                lami_calc_value = _normalize_number_text(lami_calc_value)
                if lami_calc_value is None:
                    return _ok({"detail": "Lami calc value phải là số"}, 400)
            else:
                lami_calc_value = None
            unit_weight_mode = (_str_or_none(body.get("unit_weight_mode")) or "fixed").lower()
            raw_unit_weight = _str_or_none(body.get("unit_weight_value"))
            unit_weight_value = _normalize_number_text(raw_unit_weight) if raw_unit_weight is not None else None
            unit_weight_formula_code = _str_or_none(body.get("unit_weight_formula")) or _str_or_none(
                body.get("unit_weight_formula_code")
            )
            unit_weight_option_id = body.get("unit_weight_option_id")
            unit_weight_note = _str_or_none(body.get("unit_weight_note"))
            if unit_weight_mode not in {"fixed", "formula", "choice"}:
                return _ok({"detail": "unit_weight_mode không hợp lệ"}, 400)
            if unit_weight_mode == "fixed":
                if raw_unit_weight is None or unit_weight_value is None:
                    return _ok({"detail": "Unit Weight (fixed) phải là số"}, 400)
                unit_weight_formula_code = None
                unit_weight_option_id = None
            elif unit_weight_mode == "choice":
                if not unit_weight_option_id:
                    return _ok({"detail": "Thiếu unit_weight_option_id"}, 400)
                option = session.scalar(select(UnitWeightOption).where(UnitWeightOption.id == int(unit_weight_option_id), _active(UnitWeightOption)))
                if not option:
                    return _ok({"detail": "unit_weight_option_id không hợp lệ"}, 400)
                unit_weight_value = None
                unit_weight_formula_code = None
                unit_weight_option_id = option.id
            else:
                unit_weight_formula_code = _validate_formula_expr(unit_weight_formula_code)
                if not unit_weight_formula_code:
                    return _ok({"detail": "Công thức Unit Weight không hợp lệ. Chỉ dùng A/B/C, số và + - * / ( )"}, 400)
                unit_weight_value = None
                unit_weight_option_id = None
            payload = {
                "spec_label": spec_label,
                "has_lami": has_lami,
                "use_lami_for_calc": use_lami_for_calc,
                "lami_calc_value": lami_calc_value,
                "pcs_ea_label": pcs_ea_label,
                "unit_weight_mode": unit_weight_mode,
                "unit_weight_value": unit_weight_value,
                "unit_weight_formula_code": unit_weight_formula_code,
                "unit_weight_option_id": unit_weight_option_id,
                "unit_weight_note": unit_weight_note,
            }
            existing = session.scalar(select(MaterialGroup).where(MaterialGroup.material_group_name == name))
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Material group đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.spec_label = payload["spec_label"]
                existing.has_lami = payload["has_lami"]
                existing.use_lami_for_calc = payload["use_lami_for_calc"]
                existing.lami_calc_value = payload["lami_calc_value"]
                existing.pcs_ea_label = payload["pcs_ea_label"]
                existing.unit_weight_mode = payload["unit_weight_mode"]
                existing.unit_weight_value = payload["unit_weight_value"]
                existing.unit_weight_formula_code = payload["unit_weight_formula_code"]
                existing.unit_weight_option_id = payload["unit_weight_option_id"]
                existing.unit_weight_note = payload["unit_weight_note"]
                session.add(existing)
                session.flush()
                return _ok(serialize_material_group(existing))

            item = MaterialGroup(
                material_group_name=name,
                spec_label=payload["spec_label"],
                has_lami=payload["has_lami"],
                use_lami_for_calc=payload["use_lami_for_calc"],
                lami_calc_value=payload["lami_calc_value"],
                pcs_ea_label=payload["pcs_ea_label"],
                unit_weight_mode=payload["unit_weight_mode"],
                unit_weight_value=payload["unit_weight_value"],
                unit_weight_formula_code=payload["unit_weight_formula_code"],
                unit_weight_option_id=payload["unit_weight_option_id"],
                unit_weight_note=payload["unit_weight_note"],
            )
            session.add(item)
            session.flush()
            return _ok(serialize_material_group(item), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def material_group_detail(request: HttpRequest, item_id: int):
    with get_session() as session:
        item = session.scalar(select(MaterialGroup).where(MaterialGroup.id == item_id, _active(MaterialGroup)))
        if not item:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "PUT":
            body = _body(request)
            name = _str_or_none(body.get("material_group_name"))
            if not name:
                return _ok({"detail": "Thiếu tên material group"}, 400)
            raw_spec = _str_or_none(body.get("spec_label"))
            raw_pcs = _str_or_none(body.get("pcs_ea_label"))
            spec_label = _normalize_spec_abc(raw_spec)
            pcs_ea_label = _normalize_number_text(raw_pcs) if raw_pcs is not None else (_str_or_none(item.pcs_ea_label) or "1")
            if raw_spec and not spec_label:
                return _ok({"detail": "Spec phải đúng định dạng A*B*C (A là text hoặc số, B và C là số)"}, 400)
            if raw_pcs and pcs_ea_label is None:
                return _ok({"detail": "PCS (EA) phải là số"}, 400)
            has_lami = bool(body.get("has_lami"))
            use_lami_for_calc = bool(body.get("use_lami_for_calc")) and has_lami
            lami_calc_value = _str_or_none(body.get("lami_calc_value"))
            if use_lami_for_calc:
                lami_calc_value = _normalize_number_text(lami_calc_value)
                if lami_calc_value is None:
                    return _ok({"detail": "Lami calc value phải là số"}, 400)
            else:
                lami_calc_value = None
            unit_weight_mode = (_str_or_none(body.get("unit_weight_mode")) or "fixed").lower()
            raw_unit_weight = _str_or_none(body.get("unit_weight_value"))
            unit_weight_value = _normalize_number_text(raw_unit_weight) if raw_unit_weight is not None else None
            unit_weight_formula_code = _str_or_none(body.get("unit_weight_formula")) or _str_or_none(
                body.get("unit_weight_formula_code")
            )
            unit_weight_option_id = body.get("unit_weight_option_id")
            unit_weight_note = _str_or_none(body.get("unit_weight_note"))
            if unit_weight_mode not in {"fixed", "formula", "choice"}:
                return _ok({"detail": "unit_weight_mode không hợp lệ"}, 400)
            if unit_weight_mode == "fixed":
                if raw_unit_weight is None or unit_weight_value is None:
                    return _ok({"detail": "Unit Weight (fixed) phải là số"}, 400)
                unit_weight_formula_code = None
                unit_weight_option_id = None
            elif unit_weight_mode == "choice":
                if not unit_weight_option_id:
                    return _ok({"detail": "Thiếu unit_weight_option_id"}, 400)
                option = session.scalar(select(UnitWeightOption).where(UnitWeightOption.id == int(unit_weight_option_id), _active(UnitWeightOption)))
                if not option:
                    return _ok({"detail": "unit_weight_option_id không hợp lệ"}, 400)
                unit_weight_value = None
                unit_weight_formula_code = None
                unit_weight_option_id = option.id
            else:
                unit_weight_formula_code = _validate_formula_expr(unit_weight_formula_code)
                if not unit_weight_formula_code:
                    return _ok({"detail": "Công thức Unit Weight không hợp lệ. Chỉ dùng A/B/C, số và + - * / ( )"}, 400)
                unit_weight_value = None
                unit_weight_option_id = None

            dup = session.scalar(
                select(MaterialGroup).where(
                    MaterialGroup.material_group_name == name,
                    MaterialGroup.id != item.id,
                    _active(MaterialGroup),
                )
            )
            if dup:
                return _ok({"detail": "Material group đã tồn tại"}, 400)

            item.material_group_name = name
            item.spec_label = spec_label
            item.has_lami = has_lami
            item.use_lami_for_calc = use_lami_for_calc
            item.lami_calc_value = lami_calc_value
            item.pcs_ea_label = pcs_ea_label
            item.unit_weight_mode = unit_weight_mode
            item.unit_weight_value = unit_weight_value
            item.unit_weight_formula_code = unit_weight_formula_code
            item.unit_weight_option_id = unit_weight_option_id
            item.unit_weight_note = unit_weight_note
            session.add(item)
            session.flush()
            return _ok(serialize_material_group(item))

        if request.method == "DELETE":
            in_use = session.scalar(
                select(func.count())
                .select_from(ProductSpec)
                .where(ProductSpec.material_group_id == item.id, _active(ProductSpec))
            )
            if in_use and int(in_use) > 0:
                return _ok({"detail": "Material group đang được sử dụng trong Product Specs"}, 400)
            item.deleted_at = _now()
            session.add(item)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def items(request: HttpRequest):
    _ensure_items_table_schema()
    _ensure_product_types_table()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip()
            product_type_name = (request.GET.get("product_type_name") or "").strip().upper()
            q = select(Item).where(_active(Item))
            if search:
                like = f"%{search}%"
                q = q.where(Item.item_name.ilike(like))
            if product_type_name:
                q = q.join(ItemProductType, ItemProductType.item_id == Item.id).join(
                    ProductType, ProductType.id == ItemProductType.product_type_id
                ).where(
                    ProductType.product_type_name == product_type_name,
                    _active(ProductType),
                )
            rows = session.scalars(q.order_by(Item.item_name.asc())).all()
            item_type_map = _get_item_product_type_map(session, [r.id for r in rows])
            return _ok([serialize_item(r, item_type_map) for r in rows])

        if request.method == "POST":
            body = _body(request)
            item_name = _str_or_none(body.get("item_name"))
            material_id_raw = body.get("material_id")
            item_size_source_field = (_str_or_none(body.get("item_size_source_field")) or "spec_inner").lower()
            product_type_ids_raw = body.get("product_type_ids") or []
            if not item_name:
                return _ok({"detail": "Thiếu item_name"}, 400)
            if item_size_source_field not in {"spec_inner", "top", "bottom", "liner"}:
                return _ok({"detail": "item_size_source_field không hợp lệ"}, 400)
            material_id = None
            if material_id_raw not in (None, ""):
                try:
                    material_id = int(material_id_raw)
                except (TypeError, ValueError):
                    return _ok({"detail": "material_id không hợp lệ"}, 400)
                material = session.scalar(select(Material).where(Material.id == material_id, _active(Material)))
                if not material:
                    return _ok({"detail": "material_id không tồn tại"}, 400)
            product_type_ids: list[int] = []
            for raw in product_type_ids_raw:
                try:
                    pt_id = int(raw)
                except (TypeError, ValueError):
                    return _ok({"detail": "product_type_ids không hợp lệ"}, 400)
                if pt_id not in product_type_ids:
                    product_type_ids.append(pt_id)
            if product_type_ids:
                valid_pt_ids = set(
                    session.scalars(
                        select(ProductType.id).where(ProductType.id.in_(product_type_ids), _active(ProductType))
                    ).all()
                )
                if len(valid_pt_ids) != len(product_type_ids):
                    return _ok({"detail": "product_type_ids có giá trị không tồn tại"}, 400)
            existing = session.scalar(select(Item).where(Item.item_name == item_name))
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Item đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.item_name = item_name
                existing.material_id = material_id
                existing.item_size_source_field = item_size_source_field
                session.add(existing)
                session.flush()
                session.query(ItemProductType).filter(ItemProductType.item_id == existing.id).delete()
                for pt_id in product_type_ids:
                    session.add(ItemProductType(item_id=existing.id, product_type_id=pt_id))
                session.flush()
                map_data = _get_item_product_type_map(session, [existing.id])
                return _ok(serialize_item(existing, map_data))
            obj = Item(item_name=item_name, material_id=material_id, item_size_source_field=item_size_source_field)
            session.add(obj)
            session.flush()
            for pt_id in product_type_ids:
                session.add(ItemProductType(item_id=obj.id, product_type_id=pt_id))
            session.flush()
            map_data = _get_item_product_type_map(session, [obj.id])
            return _ok(serialize_item(obj, map_data), 201)
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def item_detail(request: HttpRequest, item_id: int):
    _ensure_items_table_schema()
    _ensure_product_types_table()
    with get_session() as session:
        item = session.scalar(select(Item).where(Item.id == item_id, _active(Item)))
        if not item:
            return _ok({"detail": "Not found"}, 404)
        if request.method == "PUT":
            body = _body(request)
            item_name = _str_or_none(body.get("item_name"))
            material_id_raw = body.get("material_id")
            item_size_source_field = (_str_or_none(body.get("item_size_source_field")) or "spec_inner").lower()
            product_type_ids_raw = body.get("product_type_ids") or []
            if not item_name:
                return _ok({"detail": "Thiếu item_name"}, 400)
            if item_size_source_field not in {"spec_inner", "top", "bottom", "liner"}:
                return _ok({"detail": "item_size_source_field không hợp lệ"}, 400)
            material_id = None
            if material_id_raw not in (None, ""):
                try:
                    material_id = int(material_id_raw)
                except (TypeError, ValueError):
                    return _ok({"detail": "material_id không hợp lệ"}, 400)
                material = session.scalar(select(Material).where(Material.id == material_id, _active(Material)))
                if not material:
                    return _ok({"detail": "material_id không tồn tại"}, 400)
            product_type_ids: list[int] = []
            for raw in product_type_ids_raw:
                try:
                    pt_id = int(raw)
                except (TypeError, ValueError):
                    return _ok({"detail": "product_type_ids không hợp lệ"}, 400)
                if pt_id not in product_type_ids:
                    product_type_ids.append(pt_id)
            if product_type_ids:
                valid_pt_ids = set(
                    session.scalars(
                        select(ProductType.id).where(ProductType.id.in_(product_type_ids), _active(ProductType))
                    ).all()
                )
                if len(valid_pt_ids) != len(product_type_ids):
                    return _ok({"detail": "product_type_ids có giá trị không tồn tại"}, 400)
            dup = session.scalar(select(Item).where(Item.item_name == item_name, Item.id != item.id, _active(Item)))
            if dup:
                return _ok({"detail": "Item đã tồn tại"}, 400)
            item.item_name = item_name
            item.material_id = material_id
            item.item_size_source_field = item_size_source_field
            session.add(item)
            session.flush()
            session.query(ItemProductType).filter(ItemProductType.item_id == item.id).delete()
            for pt_id in product_type_ids:
                session.add(ItemProductType(item_id=item.id, product_type_id=pt_id))
            session.flush()
            map_data = _get_item_product_type_map(session, [item.id])
            return _ok(serialize_item(item, map_data))
        if request.method == "DELETE":
            item.deleted_at = _now()
            session.add(item)
            session.flush()
            return _ok({"success": True})
    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def raw_material_prices(request: HttpRequest):
    _ensure_raw_material_prices_table()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip()
            q = select(RawMaterialPrice).where(_active(RawMaterialPrice))
            if search:
                q = q.where(RawMaterialPrice.material_name.ilike(f"%{search}%"))
            rows = session.scalars(q.order_by(RawMaterialPrice.material_name.asc())).all()
            return _ok([serialize_raw_material_price(r) for r in rows])

        if request.method == "POST":
            body = _body(request)
            material_name = _str_or_none(body.get("material_name"))
            unit = _str_or_none(body.get("unit")) or "kg"
            unit_price = _normalize_number_text(body.get("unit_price"))
            if not material_name:
                return _ok({"detail": "Thiếu material_name"}, 400)
            if unit_price is None:
                return _ok({"detail": "Đơn giá phải là số"}, 400)

            existing = session.scalar(select(RawMaterialPrice).where(RawMaterialPrice.material_name == material_name))
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Nguyên liệu đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.unit = unit
                existing.unit_price = unit_price
                session.add(existing)
                session.flush()
                return _ok(serialize_raw_material_price(existing))

            row = RawMaterialPrice(material_name=material_name, unit=unit, unit_price=unit_price)
            session.add(row)
            session.flush()
            return _ok(serialize_raw_material_price(row), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def raw_material_price_detail(request: HttpRequest, item_id: int):
    _ensure_raw_material_prices_table()
    with get_session() as session:
        row = session.scalar(select(RawMaterialPrice).where(RawMaterialPrice.id == item_id, _active(RawMaterialPrice)))
        if not row:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "PUT":
            body = _body(request)
            material_name = _str_or_none(body.get("material_name"))
            unit = _str_or_none(body.get("unit")) or "kg"
            unit_price = _normalize_number_text(body.get("unit_price"))
            if not material_name:
                return _ok({"detail": "Thiếu material_name"}, 400)
            if unit_price is None:
                return _ok({"detail": "Đơn giá phải là số"}, 400)

            dup = session.scalar(
                select(RawMaterialPrice).where(
                    RawMaterialPrice.material_name == material_name,
                    RawMaterialPrice.id != row.id,
                    _active(RawMaterialPrice),
                )
            )
            if dup:
                return _ok({"detail": "Nguyên liệu đã tồn tại"}, 400)

            row.material_name = material_name
            row.unit = unit
            row.unit_price = unit_price
            session.add(row)
            session.flush()
            return _ok(serialize_raw_material_price(row))

        if request.method == "DELETE":
            row.deleted_at = _now()
            session.add(row)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def processing_prices(request: HttpRequest):
    _ensure_processing_prices_table()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip()
            q = select(ProcessingPrice).where(_active(ProcessingPrice))
            if search:
                q = q.where(ProcessingPrice.process_name.ilike(f"%{search}%"))
            rows = session.scalars(q.order_by(ProcessingPrice.process_name.asc())).all()
            return _ok([serialize_processing_price(r) for r in rows])

        if request.method == "POST":
            body = _body(request)
            process_name = _str_or_none(body.get("process_name"))
            unit_price = _normalize_number_text(body.get("unit_price"))
            note = _str_or_none(body.get("note"))
            if not process_name:
                return _ok({"detail": "Thiếu process_name"}, 400)
            if unit_price is None:
                return _ok({"detail": "Đơn giá phải là số"}, 400)

            existing = session.scalar(select(ProcessingPrice).where(ProcessingPrice.process_name == process_name))
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Hạng mục gia công đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.unit_price = unit_price
                existing.note = note
                session.add(existing)
                session.flush()
                return _ok(serialize_processing_price(existing))

            row = ProcessingPrice(process_name=process_name, unit_price=unit_price, note=note)
            session.add(row)
            session.flush()
            return _ok(serialize_processing_price(row), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def processing_price_detail(request: HttpRequest, item_id: int):
    _ensure_processing_prices_table()
    with get_session() as session:
        row = session.scalar(select(ProcessingPrice).where(ProcessingPrice.id == item_id, _active(ProcessingPrice)))
        if not row:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "PUT":
            body = _body(request)
            process_name = _str_or_none(body.get("process_name"))
            unit_price = _normalize_number_text(body.get("unit_price"))
            note = _str_or_none(body.get("note"))
            if not process_name:
                return _ok({"detail": "Thiếu process_name"}, 400)
            if unit_price is None:
                return _ok({"detail": "Đơn giá phải là số"}, 400)

            dup = session.scalar(
                select(ProcessingPrice).where(
                    ProcessingPrice.process_name == process_name,
                    ProcessingPrice.id != row.id,
                    _active(ProcessingPrice),
                )
            )
            if dup:
                return _ok({"detail": "Hạng mục gia công đã tồn tại"}, 400)

            row.process_name = process_name
            row.unit_price = unit_price
            row.note = note
            session.add(row)
            session.flush()
            return _ok(serialize_processing_price(row))

        if request.method == "DELETE":
            row.deleted_at = _now()
            session.add(row)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def quotation_preview(request: HttpRequest):
    _ensure_quotations_table()
    _ensure_raw_material_prices_table()
    _ensure_processing_prices_table()
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)
    with get_session() as session:
        body = _body(request)
        customer_id = body.get("customer_id")
        product_id = body.get("product_id")
        if not customer_id or not product_id:
            return _ok({"detail": "Thiếu customer_id hoặc product_id"}, 400)
        customer = session.scalar(select(Customer).where(Customer.id == int(customer_id), _active(Customer)))
        product = session.scalar(select(Product).where(Product.id == int(product_id), _active(Product)))
        if not customer or not product:
            return _ok({"detail": "Khách hàng hoặc sản phẩm không hợp lệ"}, 400)
        snapshot = _build_quotation_snapshot(
            session,
            customer,
            product,
            bool(body.get("has_lami")),
            body.get("extra_rows"),
        )
        try:
            snapshot["row_payload"] = json.loads(snapshot.get("row_payload") or "{}")
        except Exception:
            snapshot["row_payload"] = {}
        return _ok(snapshot)


def _serialize_quotation_with_refs(session, item: Quotation):
    data = serialize_quotation(item)
    customer = session.scalar(select(Customer).where(Customer.id == item.customer_id))
    product = session.scalar(select(Product).where(Product.id == item.product_id))
    data["customer_code"] = customer.customer_code if customer else None
    data["customer_name"] = customer.customer_name if customer else None
    data["product_code"] = product.product_code if product else None
    data["product_name"] = product.product_name if product else None
    return data


@csrf_exempt
@require_auth
def quotations(request: HttpRequest):
    _ensure_quotations_table()
    _ensure_raw_material_prices_table()
    _ensure_processing_prices_table()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip().lower()
            rows = session.scalars(select(Quotation).where(_active(Quotation)).order_by(Quotation.id.desc())).all()
            out = []
            for row in rows:
                item = _serialize_quotation_with_refs(session, row)
                if search:
                    hay = " ".join(
                        [
                            (item.get("customer_code") or ""),
                            (item.get("customer_name") or ""),
                            (item.get("product_code") or ""),
                            (item.get("product_name") or ""),
                        ]
                    ).lower()
                    if search not in hay:
                        continue
                out.append(item)
            return _ok(out)

        if request.method == "POST":
            body = _body(request)
            customer_id = body.get("customer_id")
            product_id = body.get("product_id")
            if not customer_id or not product_id:
                return _ok({"detail": "Thiếu customer_id hoặc product_id"}, 400)
            customer = session.scalar(select(Customer).where(Customer.id == int(customer_id), _active(Customer)))
            product = session.scalar(select(Product).where(Product.id == int(product_id), _active(Product)))
            if not customer or not product:
                return _ok({"detail": "Khách hàng hoặc sản phẩm không hợp lệ"}, 400)
            snapshot = _build_quotation_snapshot(
                session,
                customer,
                product,
                bool(body.get("has_lami")),
                body.get("extra_rows"),
            )
            row = Quotation(
                customer_id=customer.id,
                product_id=product.id,
                has_lami=snapshot["has_lami"],
                lami_unit_price=snapshot["lami_unit_price"],
                level_code=snapshot["level_code"],
                level_factor=snapshot["level_factor"],
                size_value=snapshot["size_value"],
                total_weight_kg=snapshot["total_weight_kg"],
                pe_weight_kg=snapshot["pe_weight_kg"],
                pp_weight_kg=snapshot["pp_weight_kg"],
                amount_weight=snapshot["amount_weight"],
                amount_lami=snapshot["amount_lami"],
                amount_color=snapshot["amount_color"],
                amount_extra=snapshot["amount_extra"],
                subtotal=snapshot["subtotal"],
                total=snapshot["total"],
                row_payload=snapshot["row_payload"],
                note=_str_or_none(body.get("note")),
            )
            session.add(row)
            session.flush()
            return _ok(_serialize_quotation_with_refs(session, row), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def quotation_detail(request: HttpRequest, item_id: int):
    _ensure_quotations_table()
    _ensure_raw_material_prices_table()
    _ensure_processing_prices_table()
    with get_session() as session:
        row = session.scalar(select(Quotation).where(Quotation.id == item_id, _active(Quotation)))
        if not row:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "GET":
            return _ok(_serialize_quotation_with_refs(session, row))

        if request.method == "PUT":
            body = _body(request)
            customer_id = body.get("customer_id", row.customer_id)
            product_id = body.get("product_id", row.product_id)
            customer = session.scalar(select(Customer).where(Customer.id == int(customer_id), _active(Customer)))
            product = session.scalar(select(Product).where(Product.id == int(product_id), _active(Product)))
            if not customer or not product:
                return _ok({"detail": "Khách hàng hoặc sản phẩm không hợp lệ"}, 400)
            snapshot = _build_quotation_snapshot(
                session,
                customer,
                product,
                bool(body.get("has_lami")),
                body.get("extra_rows"),
            )
            row.customer_id = customer.id
            row.product_id = product.id
            row.has_lami = snapshot["has_lami"]
            row.lami_unit_price = snapshot["lami_unit_price"]
            row.level_code = snapshot["level_code"]
            row.level_factor = snapshot["level_factor"]
            row.size_value = snapshot["size_value"]
            row.total_weight_kg = snapshot["total_weight_kg"]
            row.pe_weight_kg = snapshot["pe_weight_kg"]
            row.pp_weight_kg = snapshot["pp_weight_kg"]
            row.amount_weight = snapshot["amount_weight"]
            row.amount_lami = snapshot["amount_lami"]
            row.amount_color = snapshot["amount_color"]
            row.amount_extra = snapshot["amount_extra"]
            row.subtotal = snapshot["subtotal"]
            row.total = snapshot["total"]
            row.row_payload = snapshot["row_payload"]
            row.note = _str_or_none(body.get("note"))
            session.add(row)
            session.flush()
            return _ok(_serialize_quotation_with_refs(session, row))

        if request.method == "DELETE":
            row.deleted_at = _now()
            session.add(row)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def material_groups_import_excel(request: HttpRequest):
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)
    excel_file = request.FILES.get("file")
    if not excel_file:
        return _ok({"detail": "Thiếu file Excel"}, 400)
    if not excel_file.name.lower().endswith(".xlsx"):
        return _ok({"detail": "Chỉ hỗ trợ file .xlsx"}, 400)

    tmp_dir = Path(settings.MEDIA_ROOT) / "tmp_import"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"{uuid.uuid4().hex}.xlsx"
    with tmp_path.open("wb+") as dst:
        for chunk in excel_file.chunks():
            dst.write(chunk)

    try:
        rows = _xlsx_sheet_rows(str(tmp_path), "Item")
    except Exception:
        tmp_path.unlink(missing_ok=True)
        return _ok({"detail": "Không đọc được sheet Item"}, 400)
    finally:
        tmp_path.unlink(missing_ok=True)

    created = 0
    updated = 0
    skipped = 0
    with get_session() as session:
        for row in rows[1:]:
            if len(row) < 10:
                skipped += 1
                continue
            mg_name = _str_or_none(row[9] if len(row) > 9 else None)
            if not mg_name:
                skipped += 1
                continue
            spec_label = _str_or_none(row[10] if len(row) > 10 else None)
            pcs_ea_label = _str_or_none(row[16] if len(row) > 16 else None)

            existing = session.scalar(select(MaterialGroup).where(MaterialGroup.material_group_name == mg_name))
            if existing:
                existing.deleted_at = None
                existing.spec_label = spec_label
                existing.pcs_ea_label = pcs_ea_label
                if not existing.unit_weight_mode:
                    existing.unit_weight_mode = "fixed"
                if existing.unit_weight_mode == "fixed" and existing.unit_weight_value is None:
                    existing.unit_weight_value = 0
                session.add(existing)
                updated += 1
            else:
                session.add(
                    MaterialGroup(
                        material_group_name=mg_name,
                        spec_label=spec_label,
                        pcs_ea_label=pcs_ea_label,
                        unit_weight_mode="fixed",
                        unit_weight_value=0,
                    )
                )
                created += 1
        session.flush()

    return _ok({"success": True, "created": created, "updated": updated, "skipped": skipped})


@csrf_exempt
@require_auth
def products(request: HttpRequest):
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip()
            q = select(Product).where(_active(Product))
            if search:
                like = f"%{search}%"
                q = q.where((Product.product_code.ilike(like)) | (Product.product_name.ilike(like)))
            customer_id = request.GET.get("customer_id")
            if customer_id:
                q = q.where(Product.customer_id == int(customer_id))
            rows = session.scalars(q.order_by(Product.id.desc())).all()
            return _ok([serialize_product(r) for r in rows])

        if request.method == "POST":
            body = _body(request)
            customer = session.scalar(select(Customer).where(Customer.id == body["customer_id"], _active(Customer)))
            if not customer:
                return _ok({"detail": "Customer không tồn tại hoặc đã xóa"}, 400)
            product_code = _str_or_none(body.get("product_code"))
            if not product_code:
                return _ok({"detail": "Thiếu product_code"}, 400)
            existing = session.scalar(select(Product).where(Product.product_code == product_code))
            if existing and existing.deleted_at is None:
                return _ok({"detail": "Mã sản phẩm đã tồn tại"}, 400)
            if existing and existing.deleted_at is not None:
                existing.deleted_at = None
                existing.customer_id = body["customer_id"]
                existing.product_code = product_code
                existing.product_name = body["product_name"]
                existing.swl = body.get("swl")
                existing.type = _upper_or_none(body.get("type"))
                existing.sewing_type = _upper_or_none(body.get("sewing_type"))
                existing.print = _norm_text(body.get("print"))
                existing.spec_other = body.get("spec_other")
                existing.spec_inner = body.get("spec_inner")
                existing.color = body.get("color")
                existing.liner = body.get("liner")
                existing.top = _normalize_diameter_value(body.get("top"))
                existing.bottom = _normalize_diameter_value(body.get("bottom"))
                existing.packing = body.get("packing")
                existing.other_note = body.get("other_note")
                session.add(existing)
                session.flush()
                return _ok(serialize_product(existing), 201)
            item = Product(
                customer_id=body["customer_id"],
                product_code=product_code,
                product_name=body["product_name"],
                swl=body.get("swl"),
                type=_upper_or_none(body.get("type")),
                sewing_type=_upper_or_none(body.get("sewing_type")),
                print=_norm_text(body.get("print")),
                spec_other=body.get("spec_other"),
                spec_inner=body.get("spec_inner"),
                color=body.get("color"),
                liner=body.get("liner"),
                top=_normalize_diameter_value(body.get("top")),
                bottom=_normalize_diameter_value(body.get("bottom")),
                packing=body.get("packing"),
                other_note=body.get("other_note"),
            )
            session.add(item)
            try:
                session.flush()
            except IntegrityError:
                return _ok({"detail": "Mã sản phẩm đã tồn tại"}, 400)
            return _ok(serialize_product(item), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def products_import_excel(request: HttpRequest):
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)

    excel_file = request.FILES.get("file")
    if not excel_file:
        return _ok({"detail": "Thiếu file Excel"}, 400)
    if not excel_file.name.lower().endswith(".xlsx"):
        return _ok({"detail": "Chỉ hỗ trợ file .xlsx"}, 400)

    tmp_dir = Path(settings.MEDIA_ROOT) / "tmp_import"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"{uuid.uuid4().hex}.xlsx"
    with tmp_path.open("wb+") as dst:
        for chunk in excel_file.chunks():
            dst.write(chunk)

    try:
        rows = _xlsx_sheet_rows(str(tmp_path), "Products")
    except Exception:
        tmp_path.unlink(missing_ok=True)
        return _ok({"detail": "Không đọc được sheet Products"}, 400)
    finally:
        tmp_path.unlink(missing_ok=True)

    if not rows:
        return _ok({"detail": "Sheet Products không có dữ liệu"}, 400)

    headers = [(_str_or_none(v) or "") for v in rows[0]]
    idx = {h.strip().lower(): i for i, h in enumerate(headers) if h}
    required = {"customercode", "productname", "productcode"}
    if not required.issubset(set(idx.keys())):
        return _ok({"detail": "Thiếu cột bắt buộc CustomerCode/Productname/Productcode"}, 400)

    def val(row_values: list[str | None], key: str):
        pos = idx.get(key)
        if pos is None or pos >= len(row_values):
            return None
        return row_values[pos]

    created = 0
    skipped = 0
    failed: list[dict] = []
    with get_session() as session:
        customers = session.scalars(select(Customer).where(_active(Customer))).all()
        customer_map = {_norm_text(c.customer_code): c for c in customers}
        existing_codes = {
            _norm_text(code)
            for code in session.scalars(select(Product.product_code).where(_active(Product))).all()
            if code
        }
        seen_codes: set[str] = set()

        for line_no, row in enumerate(rows[1:], start=2):
            customer_code = _str_or_none(val(row, "customercode"))
            product_name = _str_or_none(val(row, "productname"))
            product_code = _str_or_none(val(row, "productcode"))

            if not customer_code and not product_name and not product_code:
                skipped += 1
                continue

            reasons: list[str] = []
            if not customer_code:
                reasons.append("Thiếu mã khách hàng")
            if not product_name:
                reasons.append("Thiếu tên sản phẩm")
            if not product_code:
                reasons.append("Thiếu mã sản phẩm")

            customer = customer_map.get(_norm_text(customer_code))
            if customer_code and not customer:
                reasons.append("Không tìm thấy CustomerCode")

            code_key = _norm_text(product_code)
            if code_key and (code_key in existing_codes or code_key in seen_codes):
                reasons.append("Trùng mã sản phẩm")

            if reasons:
                failed.append(
                    {
                        "row": line_no,
                        "customer_code": customer_code,
                        "product_code": product_code,
                        "product_name": product_name,
                        "reasons": reasons,
                    }
                )
                continue

            session.add(
                Product(
                    customer_id=customer.id,
                    product_code=product_code,
                    product_name=product_name,
                    swl=_str_or_none(val(row, "s.w.l")),
                    type=_upper_or_none(val(row, "type")),
                    sewing_type=_upper_or_none(val(row, "sewingtype")),
                    print=_norm_text(val(row, "print")),
                    spec_other=_str_or_none(val(row, "specother")),
                    spec_inner=_str_or_none(val(row, "specinner")),
                    color=_str_or_none(val(row, "color")),
                    liner=_str_or_none(val(row, "liner")),
                    top=_normalize_diameter_value(val(row, "top")),
                    bottom=_normalize_diameter_value(val(row, "bottom")),
                    packing=_str_or_none(val(row, "packing")),
                    other_note=_str_or_none(val(row, "other")),
                )
            )
            created += 1
            seen_codes.add(code_key)

        session.flush()

    return _ok(
        {
            "success": True,
            "created": created,
            "skipped": skipped,
            "failed_count": len(failed),
            "failed": failed[:200],
        }
    )


@csrf_exempt
@require_auth
def product_specs_import_excel(request: HttpRequest, product_id: int):
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)

    excel_file = request.FILES.get("file")
    if not excel_file:
        return _ok({"detail": "Thiếu file Excel"}, 400)
    if not excel_file.name.lower().endswith(".xlsx"):
        return _ok({"detail": "Chỉ hỗ trợ file .xlsx"}, 400)

    tmp_dir = Path(settings.MEDIA_ROOT) / "tmp_import"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"{uuid.uuid4().hex}.xlsx"
    with tmp_path.open("wb+") as dst:
        for chunk in excel_file.chunks():
            dst.write(chunk)

    rows = None
    try:
        for sheet_name in ["Products_S", "Product_Specs", "Product Specs", "Products S"]:
            try:
                rows = _xlsx_sheet_rows(str(tmp_path), sheet_name)
                break
            except ValueError:
                continue
    except Exception:
        tmp_path.unlink(missing_ok=True)
        return _ok({"detail": "Không đọc được file Excel"}, 400)
    finally:
        tmp_path.unlink(missing_ok=True)

    if rows is None:
        return _ok({"detail": "Không tìm thấy sheet Products_S"}, 400)
    if not rows:
        return _ok({"detail": "Sheet Products_S không có dữ liệu"}, 400)

    def norm_header(v: str | None) -> str:
        raw = _str_or_none(v) or ""
        return "".join(ch for ch in raw.lower() if ch.isalnum())

    headers = [norm_header(v) for v in rows[0]]
    idx = {h: i for i, h in enumerate(headers) if h}

    aliases = {
        "product_code": ["productcode"],
        "line_no": ["lineno", "line", "no"],
        "item_name": ["itemname", "item"],
        "material_group": ["materialgroup", "materialgroupname"],
        "spec": ["spec"],
        "item_size": ["itemsize"],
        "lami": ["lami"],
        "item_color": ["itemcolor"],
        "unit_weight_kg": ["unitweightkg", "unitweight"],
        "qty_m_or_m2": ["qtymorm2", "qty"],
        "pcs_ea": ["pcsea", "pcs"],
        "wt_kg": ["wtkg", "wt"],
        "other_note": ["other", "othernote"],
    }

    def header_index(name: str) -> int | None:
        for key in aliases[name]:
            if key in idx:
                return idx[key]
        return None

    if header_index("item_name") is None or header_index("material_group") is None:
        return _ok({"detail": "Thiếu cột bắt buộc Item/MaterialGroup"}, 400)

    def val(row_values: list[str | None], name: str):
        pos = header_index(name)
        if pos is None or pos >= len(row_values):
            return None
        return row_values[pos]

    def num_or_none(v):
        raw = _str_or_none(v)
        if raw is None:
            return None
        try:
            return float(raw.replace(",", ""))
        except ValueError:
            return None

    with get_session() as session:
        product = session.scalar(select(Product).where(Product.id == product_id, _active(Product)))
        if not product:
            return _ok({"detail": "Product không tồn tại hoặc đã xóa"}, 404)

        max_line_no = session.scalar(
            select(func.max(ProductSpec.line_no)).where(ProductSpec.product_id == product_id, _active(ProductSpec))
        ) or 0
        existing_line_nos = set(
            session.scalars(
                select(ProductSpec.line_no).where(ProductSpec.product_id == product_id, _active(ProductSpec))
            ).all()
        )

        created = 0
        skipped = 0
        failed: list[dict] = []
        seen_line_nos: set[int] = set()

        for line_no, row in enumerate(rows[1:], start=2):
            product_code = _str_or_none(val(row, "product_code"))
            item_name = _str_or_none(val(row, "item_name"))
            material_group_name = _str_or_none(val(row, "material_group"))
            line_value = _str_or_none(val(row, "line_no"))

            if not product_code and not item_name and not material_group_name and not line_value:
                skipped += 1
                continue

            reasons: list[str] = []

            if product_code and _norm_text(product_code) != _norm_text(product.product_code):
                reasons.append("ProductCode không khớp sản phẩm đang chọn")
            if not item_name:
                reasons.append("Thiếu Item")
            if not material_group_name:
                reasons.append("Thiếu MaterialGroup")

            if line_value:
                try:
                    spec_line_no = int(float(line_value))
                    if spec_line_no <= 0:
                        reasons.append("LineNo phải lớn hơn 0")
                except ValueError:
                    spec_line_no = None
                    reasons.append("LineNo không hợp lệ")
            else:
                max_line_no += 1
                spec_line_no = max_line_no

            if spec_line_no and (spec_line_no in existing_line_nos or spec_line_no in seen_line_nos):
                reasons.append("Trùng LineNo trong sản phẩm")

            unit_weight_kg = num_or_none(val(row, "unit_weight_kg"))
            qty_m_or_m2 = num_or_none(val(row, "qty_m_or_m2"))
            pcs_ea = num_or_none(val(row, "pcs_ea"))
            wt_kg = num_or_none(val(row, "wt_kg"))

            if _str_or_none(val(row, "unit_weight_kg")) and unit_weight_kg is None:
                reasons.append("Unit Weight(kg) không hợp lệ")
            if _str_or_none(val(row, "qty_m_or_m2")) and qty_m_or_m2 is None:
                reasons.append("Q'ty(m or m2) không hợp lệ")
            if _str_or_none(val(row, "pcs_ea")) and pcs_ea is None:
                reasons.append("PCS(EA) không hợp lệ")
            if _str_or_none(val(row, "wt_kg")) and wt_kg is None:
                reasons.append("WT (kg) không hợp lệ")

            if reasons:
                failed.append(
                    {
                        "row": line_no,
                        "product_code": product_code or product.product_code,
                        "line_no": spec_line_no,
                        "item_name": item_name,
                        "material_group": material_group_name,
                        "reasons": reasons,
                    }
                )
                continue

            item = _get_or_create_item(session, item_name)
            material_group = _get_or_create_material_group(session, material_group_name)
            if not item or not material_group or not spec_line_no:
                failed.append(
                    {
                        "row": line_no,
                        "product_code": product_code or product.product_code,
                        "line_no": spec_line_no,
                        "item_name": item_name,
                        "material_group": material_group_name,
                        "reasons": ["Không tạo được Item hoặc MaterialGroup"],
                    }
                )
                continue

            session.add(
                ProductSpec(
                    product_id=product_id,
                    item_id=item.id,
                    material_group_id=material_group.id,
                    line_no=spec_line_no,
                    spec=_str_or_none(val(row, "spec")),
                    item_size=_str_or_none(val(row, "item_size")),
                    lami=_str_or_none(val(row, "lami")),
                    item_color=_str_or_none(val(row, "item_color")),
                    unit_weight_kg=unit_weight_kg,
                    qty_m_or_m2=qty_m_or_m2,
                    pcs_ea=pcs_ea,
                    wt_kg=wt_kg,
                    other_note=_str_or_none(val(row, "other_note")),
                    is_manual_weight=True,
                )
            )
            created += 1
            seen_line_nos.add(spec_line_no)

        session.flush()

    return _ok(
        {
            "success": True,
            "created": created,
            "skipped": skipped,
            "failed_count": len(failed),
            "failed": failed[:200],
        }
    )


@csrf_exempt
@require_auth
def product_detail(request: HttpRequest, item_id: int):
    _ensure_items_table_schema()
    _ensure_product_types_table()
    with get_session() as session:
        item = session.scalar(select(Product).where(Product.id == item_id, _active(Product)))
        if not item:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "GET":
            return _ok(serialize_product(item))

        if request.method == "PUT":
            body = _body(request)
            if "customer_id" in body:
                customer = session.scalar(select(Customer).where(Customer.id == body["customer_id"], _active(Customer)))
                if not customer:
                    return _ok({"detail": "Customer không tồn tại hoặc đã xóa"}, 400)
            original_type = (item.type or "").strip().upper()
            need_recompute_spec_qty = False
            for f in [
                "customer_id",
                "product_code",
                "product_name",
                "swl",
                "type",
                "sewing_type",
                "print",
                "spec_other",
                "spec_inner",
                "color",
                "liner",
                "top",
                "bottom",
                "packing",
                "other_note",
            ]:
                if f in body:
                    if f in {"type", "sewing_type"}:
                        setattr(item, f, _upper_or_none(body[f]))
                    elif f == "print":
                        setattr(item, f, _norm_text(body[f]))
                    elif f in {"top", "bottom"}:
                        setattr(item, f, _normalize_diameter_value(body[f]))
                    else:
                        setattr(item, f, body[f])
                    if f in {"type", "spec_inner", "top", "bottom", "liner"}:
                        need_recompute_spec_qty = True
            if "product_code" in body:
                next_code = _str_or_none(body.get("product_code"))
                if not next_code:
                    return _ok({"detail": "Thiếu product_code"}, 400)
                dup_active = session.scalar(
                    select(Product).where(Product.product_code == next_code, Product.id != item.id, _active(Product))
                )
                if dup_active:
                    return _ok({"detail": "Mã sản phẩm đã tồn tại"}, 400)
                dup_deleted = session.scalar(
                    select(Product).where(Product.product_code == next_code, Product.id != item.id, Product.deleted_at.is_not(None))
                )
                if dup_deleted:
                    return _ok({"detail": "Mã sản phẩm đã tồn tại ở dữ liệu đã xóa. Vui lòng dùng mã khác."}, 400)
            if need_recompute_spec_qty:
                _recompute_product_specs_item_size_qty(session, item)
            removed_spec_rows: list[ProductSpec] = []
            next_type = (item.type or "").strip().upper()
            if next_type and next_type != original_type:
                next_pt = session.scalar(
                    select(ProductType).where(
                        ProductType.product_type_name == next_type,
                        _active(ProductType),
                    )
                )
                if next_pt:
                    specs = session.scalars(
                        select(ProductSpec).where(
                            ProductSpec.product_id == item.id,
                            _active(ProductSpec),
                        )
                    ).all()
                    if specs:
                        spec_item_ids = list({s.item_id for s in specs})
                        allowed_item_ids = set(
                            session.scalars(
                                select(ItemProductType.item_id).where(
                                    ItemProductType.product_type_id == next_pt.id,
                                    ItemProductType.item_id.in_(spec_item_ids),
                                )
                            ).all()
                        )
                        now = _now()
                        for spec in specs:
                            if spec.item_id not in allowed_item_ids:
                                spec.deleted_at = now
                                session.add(spec)
                                removed_spec_rows.append(spec)
            session.add(item)
            try:
                session.flush()
            except IntegrityError:
                return _ok({"detail": "Mã sản phẩm đã tồn tại"}, 400)
            res = serialize_product(item)
            res["removed_spec_count"] = len(removed_spec_rows)
            res["removed_spec_items"] = sorted(
                list(
                    {
                        s.item.item_name
                        for s in removed_spec_rows
                        if s.item is not None and s.item.item_name
                    }
                )
            )
            return _ok(res)

        if request.method == "DELETE":
            soft_delete_product(session, item.id)
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def product_export_excel(request: HttpRequest, product_id: int):
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)

    body = _body(request)
    mode = (_str_or_none(body.get("mode")) or "").strip().lower()
    if mode not in {"form_product", "form_specification"}:
        return _ok({"detail": "mode không hợp lệ"}, 400)

    try:
        from openpyxl import Workbook  # type: ignore
    except Exception:
        return _ok({"detail": "Thiếu thư viện openpyxl. Vui lòng cài openpyxl để xuất Excel"}, 500)

    try:
        with get_session() as session:
            product = session.scalar(select(Product).where(Product.id == product_id, _active(Product)))
            if not product:
                return _ok({"detail": "Product không tồn tại hoặc đã xóa"}, 404)
            customer = session.scalar(select(Customer).where(Customer.id == product.customer_id, _active(Customer)))

            specs = session.scalars(
                select(ProductSpec)
                .where(ProductSpec.product_id == product_id, _active(ProductSpec))
                .order_by(ProductSpec.id.asc())
            ).all()
            print_images = session.scalars(
                select(ProductPrintImage)
                .join(ProductPrintVersion, ProductPrintImage.product_print_version_id == ProductPrintVersion.id)
                .where(
                    ProductPrintVersion.product_id == product_id,
                    _active(ProductPrintVersion),
                    _active(ProductPrintImage),
                )
                .order_by(ProductPrintVersion.version_no.desc(), ProductPrintImage.sort_order.asc())
            ).all()

            if mode == "form_product":
                selected_ids = body.get("spec_ids") or []
                try:
                    selected_ids = [int(x) for x in selected_ids]
                except Exception:
                    return _ok({"detail": "spec_ids không hợp lệ"}, 400)
                if not selected_ids:
                    return _ok({"detail": "Vui lòng chọn ít nhất 1 spec để xuất form product"}, 400)
                selected_set = set(selected_ids)
                specs = [s for s in specs if s.id in selected_set]
                if not specs:
                    return _ok({"detail": "Không có spec hợp lệ để xuất"}, 400)

            wb = Workbook()
            ws = wb.active
            ws.title = "Form_Product" if mode == "form_product" else "Form_Specification"

            if mode == "form_product":
                _apply_form_product_sheet(ws, product, customer, specs, print_images)
                suffix = "form_product"
            else:
                _apply_form_specification_sheet(ws, product, customer, specs)
                suffix = "form_specification"

            out = BytesIO()
            wb.save(out)
            out.seek(0)
            filename = f"{product.product_code or 'product'}_{suffix}.xlsx"
            resp = HttpResponse(
                out.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp
    except Exception as exc:
        return _ok({"detail": f"Lỗi xuất Excel: {exc}"}, 500)


@csrf_exempt
@require_auth
def product_specs(request: HttpRequest, product_id: int):
    _ensure_product_specs_schema()
    _ensure_fixed_weight_tables_table()
    with get_session() as session:
        product = session.scalar(select(Product).where(Product.id == product_id, _active(Product)))
        if not product:
            return _ok({"detail": "Product không tồn tại hoặc đã xóa"}, 404)

        if request.method == "GET":
            rows = session.scalars(
                select(ProductSpec)
                .where(ProductSpec.product_id == product_id, _active(ProductSpec))
                .order_by(ProductSpec.line_no.asc())
            ).all()
            for row in rows:
                _sync_product_spec_from_item_material(session, row)
            session.flush()
            return _ok([serialize_spec(r) for r in rows])

        if request.method == "POST":
            body = _body(request)
            item_id = body.get("item_id")

            item = None
            if item_id:
                item = session.scalar(select(Item).where(Item.id == int(item_id), _active(Item)))
            if not item:
                item = _get_or_create_item(session, body.get("item_name"))
            if not item:
                return _ok({"detail": "Thiếu item hoặc item không hợp lệ"}, 400)
            if not item.material_id:
                return _ok({"detail": "Item chưa gán material"}, 400)

            raw_item_color = _str_or_none(body.get("item_color"))
            if raw_item_color in {"-", "--"}:
                raw_item_color = None

            input_lami = _normalize_lami_text(body.get("lami"))
            requested_spec = _str_or_none(body.get("spec"))
            if not requested_spec:
                requested_spec = _resolve_item_size_source_value(item.item_size_source_field, product, None)
            computed_unit_weight, item_material, resolved_spec = _compute_unit_weight_from_item_material(
                session,
                item,
                requested_spec,
                input_lami,
            )
            if item_material is None:
                return _ok({"detail": "Material của item không hợp lệ hoặc đã bị xóa"}, 400)
            if not resolved_spec:
                return _ok({"detail": "Thiếu spec hợp lệ cho material của item"}, 400)
            category_name = item_material.material_category.material_category_name if item_material.material_category else None
            category_code = item_material.material_category.material_category_code if item_material.material_category else None
            if _is_fabric_category(category_code, category_name):
                resolved_lami = input_lami or ("Yes" if bool(item_material.lami) else "No")
            elif _is_rope_category(category_code, category_name):
                resolved_lami = "No"
            else:
                resolved_lami = input_lami or "No"

            auto_item_size = _compute_item_size_by_product_type_formula(session, product, item, resolved_spec)
            computed_item_size = body.get("item_size") if body.get("item_size") is not None else auto_item_size
            computed_qty = (
                body.get("qty_m_or_m2")
                if body.get("qty_m_or_m2") is not None
                else _compute_qty_from_item_size(computed_item_size)
            )
            computed_wt = _compute_wt_kg(computed_unit_weight, computed_qty, body.get("pcs_ea"))

            try:
                spec_row = ProductSpec(
                    product_id=product_id,
                    item_id=item.id,
                    material_group_id=None,
                    line_no=_next_product_spec_line_no(session, product_id),
                    spec=resolved_spec,
                    item_size=computed_item_size,
                    lami=resolved_lami,
                    item_color=raw_item_color or product.color,
                    unit_weight_kg=computed_unit_weight,
                    qty_m_or_m2=computed_qty,
                    pcs_ea=body.get("pcs_ea"),
                    wt_kg=computed_wt,
                    other_note=body.get("other_note"),
                )
                session.add(spec_row)
                session.flush()
                return _ok(serialize_spec(spec_row), 201)
            except IntegrityError:
                return _ok({"detail": "Không thể tạo Product Spec do trùng LineNo, vui lòng thử lại"}, 409)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def product_spec_detail(request: HttpRequest, spec_id: int):
    _ensure_product_specs_schema()
    _ensure_fixed_weight_tables_table()
    with get_session() as session:
        item = session.scalar(select(ProductSpec).where(ProductSpec.id == spec_id, _active(ProductSpec)))
        if not item:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "PUT":
            body = _body(request)
            recompute_weight = False
            recompute_item_size = False
            next_item: Item | None = None
            if "item_id" in body or "item_name" in body:
                if body.get("item_id"):
                    next_item = session.scalar(select(Item).where(Item.id == int(body.get("item_id")), _active(Item)))
                if not next_item and "item_name" in body:
                    next_item = _get_or_create_item(session, body.get("item_name"))
                if not next_item:
                    return _ok({"detail": "item không hợp lệ"}, 400)
                item.item_id = next_item.id
                recompute_weight = True
                recompute_item_size = True

            for f in [
                "line_no",
                "spec",
                "lami",
                "item_size",
                "item_color",
                "qty_m_or_m2",
                "pcs_ea",
                "other_note",
            ]:
                if f in body:
                    setattr(item, f, body[f])
                    if f in {"spec", "lami"}:
                        recompute_weight = True
                    if f == "spec":
                        recompute_item_size = True
            current_item = next_item
            if current_item is None:
                current_item = session.scalar(select(Item).where(Item.id == item.item_id, _active(Item)))
            if recompute_item_size and "item_size" not in body:
                product = session.scalar(select(Product).where(Product.id == item.product_id, _active(Product)))
                auto_item_size = _compute_item_size_by_product_type_formula(session, product, current_item, item.spec)
                if auto_item_size:
                    item.item_size = auto_item_size
            if ("item_size" in body or recompute_item_size) and "qty_m_or_m2" not in body:
                item.qty_m_or_m2 = _compute_qty_from_item_size(item.item_size)
            if not current_item or not current_item.material_id:
                return _ok({"detail": "Item chưa gán material"}, 400)
            normalized_lami = _normalize_lami_text(item.lami)
            computed_unit_weight, item_material, resolved_spec = _compute_unit_weight_from_item_material(
                session,
                current_item,
                item.spec,
                normalized_lami,
            )
            if item_material is None:
                return _ok({"detail": "Material của item không hợp lệ hoặc đã bị xóa"}, 400)
            if not resolved_spec:
                return _ok({"detail": "Thiếu spec hợp lệ cho material của item"}, 400)
            category_name = item_material.material_category.material_category_name if item_material.material_category else None
            category_code = item_material.material_category.material_category_code if item_material.material_category else None
            if _is_fabric_category(category_code, category_name):
                item.spec = resolved_spec
                item.lami = normalized_lami or ("Yes" if bool(item_material.lami) else "No")
            elif _is_rope_category(category_code, category_name):
                item.spec = resolved_spec
                item.lami = "No"
            else:
                item.spec = resolved_spec
                item.lami = normalized_lami or "No"
            item.unit_weight_kg = computed_unit_weight
            item.wt_kg = _compute_wt_kg(item.unit_weight_kg, item.qty_m_or_m2, item.pcs_ea)
            session.add(item)
            session.flush()
            return _ok(serialize_spec(item))

        if request.method == "DELETE":
            item.deleted_at = _now()
            session.add(item)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def product_print_versions(request: HttpRequest, product_id: int):
    with get_session() as session:
        if request.method != "GET":
            return _ok({"detail": "Method not allowed"}, 405)
        product = session.scalar(select(Product).where(Product.id == product_id, _active(Product)))
        if not product:
            return _ok({"detail": "Product không tồn tại hoặc đã xóa"}, 404)

        rows = session.scalars(
            select(ProductPrintVersion)
            .where(ProductPrintVersion.product_id == product_id, _active(ProductPrintVersion))
            .order_by(ProductPrintVersion.version_no.desc())
        ).all()
        out = []
        for r in rows:
            count = session.scalar(
                select(func.count(ProductPrintImage.id)).where(
                    ProductPrintImage.product_print_version_id == r.id,
                    _active(ProductPrintImage),
                )
            )
            item = serialize_version(r)
            item["image_count"] = count
            out.append(item)
        return _ok(out)


@csrf_exempt
@require_auth
def product_print_images(request: HttpRequest, product_id: int):
    with get_session() as session:
        if request.method != "GET":
            return _ok({"detail": "Method not allowed"}, 405)
        product = session.scalar(select(Product).where(Product.id == product_id, _active(Product)))
        if not product:
            return _ok({"detail": "Product không tồn tại hoặc đã xóa"}, 404)

        rows = session.execute(
            select(ProductPrintImage, ProductPrintVersion)
            .join(ProductPrintVersion, ProductPrintImage.product_print_version_id == ProductPrintVersion.id)
            .where(
                ProductPrintVersion.product_id == product_id,
                _active(ProductPrintVersion),
                _active(ProductPrintImage),
            )
            .order_by(ProductPrintVersion.version_no.desc(), ProductPrintImage.sort_order.asc())
        ).all()
        out = []
        for image, version in rows:
            item = serialize_image(image)
            item["version_no"] = version.version_no
            item["product_id"] = version.product_id
            out.append(item)
        return _ok(out)


@csrf_exempt
@require_auth
def product_print_upload(request: HttpRequest, product_id: int):
    if request.method != "POST":
        return _ok({"detail": "Method not allowed"}, 405)

    files = request.FILES.getlist("images")
    if not files:
        return _ok({"detail": "Không có ảnh upload"}, 400)

    upload_note = request.POST.get("upload_note")
    current_user = request.current_user

    try:
        with get_session() as session:
            product = session.scalar(select(Product).where(Product.id == product_id, _active(Product)))
            if not product:
                return _ok({"detail": "Product không tồn tại hoặc đã xóa"}, 404)

            last_version_no = session.scalar(
                select(func.max(ProductPrintVersion.version_no)).where(ProductPrintVersion.product_id == product_id, _active(ProductPrintVersion))
            ) or 0

            created_versions = []
            for f in files:
                last_version_no += 1
                version = ProductPrintVersion(
                    product_id=product_id,
                    version_no=last_version_no,
                    upload_note=upload_note,
                    created_by=current_user.username,
                )
                session.add(version)
                session.flush()

                target_dir = Path(settings.MEDIA_ROOT) / "print_images" / str(product_id) / str(version.version_no)
                target_dir.mkdir(parents=True, exist_ok=True)
                ext = Path(f.name).suffix.lower()
                file_name = f"001_{uuid.uuid4().hex}{ext}"
                abs_path = target_dir / file_name
                with abs_path.open("wb+") as dst:
                    for chunk in f.chunks():
                        dst.write(chunk)
                rel_path = os.path.relpath(abs_path, settings.MEDIA_ROOT)
                image_url = request.build_absolute_uri(settings.MEDIA_URL + rel_path.replace("\\", "/"))
                img = ProductPrintImage(
                    product_print_version_id=version.id,
                    image_url=image_url,
                    file_name=f.name,
                    mime_type=f.content_type,
                    file_size=f.size,
                    sort_order=1,
                )
                session.add(img)
                created_versions.append(serialize_version(version))

            _sync_product_has_print_assets(session, product_id)
            session.flush()
            return _ok({"created": len(created_versions), "versions": created_versions}, 201)
    except Exception as exc:
        return _ok({"detail": f"Upload ảnh thất bại: {exc}"}, 500)


@csrf_exempt
@require_auth
def print_version_detail(request: HttpRequest, version_id: int):
    with get_session() as session:
        version = session.scalar(select(ProductPrintVersion).where(ProductPrintVersion.id == version_id, _active(ProductPrintVersion)))
        if not version:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "GET":
            images = session.scalars(
                select(ProductPrintImage)
                .where(ProductPrintImage.product_print_version_id == version_id, _active(ProductPrintImage))
                .order_by(ProductPrintImage.sort_order.asc())
            ).all()
            return _ok({"version": serialize_version(version), "images": [serialize_image(i) for i in images]})

        if request.method == "DELETE":
            soft_delete_print_version(session, version_id)
            _sync_product_has_print_assets(session, version.product_id)
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def print_image_detail(request: HttpRequest, image_id: int):
    with get_session() as session:
        image = session.scalar(select(ProductPrintImage).where(ProductPrintImage.id == image_id, _active(ProductPrintImage)))
        if not image:
            return _ok({"detail": "Not found"}, 404)
        version = session.scalar(
            select(ProductPrintVersion).where(
                ProductPrintVersion.id == image.product_print_version_id,
                _active(ProductPrintVersion),
            )
        )
        if not version:
            return _ok({"detail": "Version not found"}, 404)

        if request.method == "DELETE":
            image.deleted_at = _now()
            session.add(image)
            session.flush()

            remain_count = session.scalar(
                select(func.count(ProductPrintImage.id)).where(
                    ProductPrintImage.product_print_version_id == version.id,
                    _active(ProductPrintImage),
                )
            ) or 0
            if remain_count == 0:
                soft_delete_print_version(session, version.id)
            _sync_product_has_print_assets(session, version.product_id)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def production_plans(request: HttpRequest):
    _ensure_production_plan_note_column()
    with get_session() as session:
        if request.method == "GET":
            search = (request.GET.get("search") or "").strip()
            q = select(ProductionPlan).where(_active(ProductionPlan))
            if search:
                q = q.where(ProductionPlan.lot_no.ilike(f"%{search}%"))
            rows = session.scalars(q.order_by(ProductionPlan.id.desc())).all()
            return _ok([serialize_plan(r) for r in rows])

        if request.method == "POST":
            body = _body(request)
            customer = session.scalar(select(Customer).where(Customer.id == body["customer_id"], _active(Customer)))
            p = session.scalar(select(Product).where(Product.id == body["product_id"], _active(Product)))
            if not customer or not p or p.customer_id != int(body["customer_id"]):
                return _ok({"detail": "Product không thuộc Customer đã chọn"}, 400)
            item = ProductionPlan(
                customer_id=body["customer_id"],
                product_id=body["product_id"],
                lot_no=body["lot_no"],
                etd=parse_date(body.get("etd")),
                eta=parse_date(body.get("eta")),
                contp_date=parse_date(body.get("contp_date")),
                order_qty_pcs=body.get("order_qty_pcs") or 0,
                spec_inner_snapshot=_norm_text(body.get("spec_inner_snapshot")) if body.get("spec_inner_snapshot") is not None else p.spec_inner,
                liner_snapshot=_norm_text(body.get("liner_snapshot")) if body.get("liner_snapshot") is not None else p.liner,
                print_snapshot=_norm_text(body.get("print_snapshot")) if body.get("print_snapshot") is not None else p.print,
                label=_norm_text(body.get("label")) if body.get("label") is not None else p.product_name,
                sewing_type=_norm_text(body.get("sewing_type")) if body.get("sewing_type") is not None else p.sewing_type,
                packing=_norm_text(body.get("packing")) if body.get("packing") is not None else p.packing,
                note=_str_or_none(body.get("note")),
                status=body.get("status") or "draft",
                update_person=body.get("update_person"),
            )
            session.add(item)
            session.flush()
            return _ok(serialize_plan(item), 201)

    return _ok({"detail": "Method not allowed"}, 405)


@csrf_exempt
@require_auth
def production_plan_detail(request: HttpRequest, item_id: int):
    _ensure_production_plan_note_column()
    with get_session() as session:
        item = session.scalar(select(ProductionPlan).where(ProductionPlan.id == item_id, _active(ProductionPlan)))
        if not item:
            return _ok({"detail": "Not found"}, 404)

        if request.method == "PUT":
            body = _body(request)
            for f in [
                "customer_id",
                "product_id",
                "lot_no",
                "order_qty_pcs",
                "spec_inner_snapshot",
                "liner_snapshot",
                "print_snapshot",
                "label",
                "sewing_type",
                "packing",
                "note",
                "status",
                "update_person",
            ]:
                if f in body:
                    if f in {"spec_inner_snapshot", "liner_snapshot", "print_snapshot", "label", "sewing_type", "packing"}:
                        setattr(item, f, _norm_text(body[f]))
                    elif f == "note":
                        setattr(item, f, _str_or_none(body[f]))
                    else:
                        setattr(item, f, body[f])
            if "etd" in body:
                item.etd = parse_date(body.get("etd"))
            if "eta" in body:
                item.eta = parse_date(body.get("eta"))
            if "contp_date" in body:
                item.contp_date = parse_date(body.get("contp_date"))

            customer = session.scalar(select(Customer).where(Customer.id == item.customer_id, _active(Customer)))
            p = session.scalar(select(Product).where(Product.id == item.product_id, _active(Product)))
            if not customer or not p or p.customer_id != int(item.customer_id):
                return _ok({"detail": "Product không thuộc Customer đã chọn"}, 400)

            # Default from selected product only when client does not provide value.
            if "spec_inner_snapshot" not in body and not item.spec_inner_snapshot:
                item.spec_inner_snapshot = p.spec_inner
            if "liner_snapshot" not in body and not item.liner_snapshot:
                item.liner_snapshot = p.liner
            if "print_snapshot" not in body and not item.print_snapshot:
                item.print_snapshot = p.print
            if "label" not in body and not item.label:
                item.label = p.product_name
            if "sewing_type" not in body and not item.sewing_type:
                item.sewing_type = p.sewing_type
            if "packing" not in body and not item.packing:
                item.packing = p.packing

            session.add(item)
            session.flush()
            return _ok(serialize_plan(item))

        if request.method == "DELETE":
            item.deleted_at = _now()
            session.add(item)
            session.flush()
            return _ok({"success": True})

    return _ok({"detail": "Method not allowed"}, 405)
